"""Excel report renderer."""

from datetime import datetime
from analytics.report_models import Report

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelRenderer:
    """Renders reports to Excel format."""

    @staticmethod
    def render(report: Report) -> bytes:
        """Render a report to Excel bytes."""
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel rendering")

        wb = Workbook()
        wb.remove(wb.active)

        ExcelRenderer._create_summary_sheet(wb, report)
        ExcelRenderer._create_products_sheet(wb, report)
        ExcelRenderer._create_platforms_sheet(wb, report)
        ExcelRenderer._create_advertising_sheet(wb, report)
        ExcelRenderer._create_profitability_sheet(wb, report)
        ExcelRenderer._create_quality_sheet(wb, report)
        ExcelRenderer._create_insights_sheet(wb, report)
        ExcelRenderer._create_recommendations_sheet(wb, report)
        ExcelRenderer._create_metrics_sheet(wb, report)

        from io import BytesIO
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    @staticmethod
    def _get_header_style():
        """Get header cell style."""
        return {
            "font": Font(bold=True, color="FFFFFF", size=11),
            "fill": PatternFill(start_color="34495E", end_color="34495E", fill_type="solid"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        }

    @staticmethod
    def _get_title_style():
        """Get title cell style."""
        return {
            "font": Font(bold=True, size=14, color="2C3E50"),
            "alignment": Alignment(horizontal="left", vertical="center"),
        }

    @staticmethod
    def _create_summary_sheet(wb, report):
        """Create summary sheet."""
        ws = wb.create_sheet("Summary", 0)

        row = 1
        ws[f"A{row}"] = report.title
        for cell in [ws[f"A{row}"], ws[f"B{row}"]]:
            for key, val in ExcelRenderer._get_title_style().items():
                setattr(cell, key, val)

        row += 2
        ws[f"A{row}"] = "Report Information"
        for key, val in ExcelRenderer._get_title_style().items():
            setattr(ws[f"A{row}"], key, val)

        row += 1
        headers = ["Field", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            for key, val in ExcelRenderer._get_header_style().items():
                setattr(cell, key, val)

        data = [
            ("Report ID", report.report_id),
            ("Report Date", report.report_date.isoformat()),
            ("Report Type", report.report_type.value),
            ("Generated", report.generated_at.isoformat()),
            ("Data Completeness", f"{report.data_completeness_pct:.0f}%"),
        ]

        row += 1
        for field, value in data:
            ws[f"A{row}"] = field
            ws[f"B{row}"] = str(value)
            row += 1

        row += 2
        ws[f"A{row}"] = "Executive Summary"
        for key, val in ExcelRenderer._get_title_style().items():
            setattr(ws[f"A{row}"], key, val)

        row += 1
        ws[f"A{row}"] = report.executive_summary
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40

    @staticmethod
    def _create_products_sheet(wb, report):
        """Create products sheet."""
        if not report.product_sections:
            return

        ws = wb.create_sheet("Products")

        row = 1
        headers = [
            "SKU", "Product Name", "Units Sold", "Net Sales (INR)",
            "Ad Spend (INR)", "ROAS", "ACOS (%)", "Organic Share (%)",
            "Profit Margin (%)", "Status", "Return Rate (%)", "Cancellation (%)"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            for key, val in ExcelRenderer._get_header_style().items():
                setattr(cell, key, val)

        row += 1
        for product in report.product_sections:
            ws[f"A{row}"] = product.sku
            ws[f"B{row}"] = product.product_name
            ws[f"C{row}"] = product.units_sold
            ws[f"D{row}"] = product.net_sales_inr
            ws[f"E{row}"] = product.ad_spend_inr
            ws[f"F{row}"] = product.roas
            ws[f"G{row}"] = product.acos_pct
            ws[f"H{row}"] = product.organic_share_pct
            ws[f"I{row}"] = product.profit_margin_pct
            ws[f"J{row}"] = product.profitability_status
            ws[f"K{row}"] = product.return_rate_pct
            ws[f"L{row}"] = product.cancellation_rate_pct
            row += 1

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + i)].width = 15

    @staticmethod
    def _create_platforms_sheet(wb, report):
        """Create platforms sheet."""
        if not report.platform_sections:
            return

        ws = wb.create_sheet("Platforms")

        row = 1
        headers = [
            "Platform ID", "Platform Name", "Product Count", "Total Units",
            "Total Sales (INR)", "Ad Spend (INR)", "ROAS", "ACOS (%)",
            "Organic Sales (INR)", "Organic Share (%)", "Profit Margin (%)"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            for key, val in ExcelRenderer._get_header_style().items():
                setattr(cell, key, val)

        row += 1
        for platform in report.platform_sections:
            ws[f"A{row}"] = platform.platform_id
            ws[f"B{row}"] = platform.platform_name
            ws[f"C{row}"] = platform.product_count
            ws[f"D{row}"] = platform.total_units_sold
            ws[f"E{row}"] = platform.total_net_sales_inr
            ws[f"F{row}"] = platform.total_ad_spend_inr
            ws[f"G{row}"] = platform.platform_roas
            ws[f"H{row}"] = platform.platform_acos_pct
            ws[f"I{row}"] = platform.total_organic_sales_inr
            ws[f"J{row}"] = platform.organic_share_pct
            ws[f"K{row}"] = platform.overall_profit_margin_pct
            row += 1

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + i)].width = 15

    @staticmethod
    def _create_advertising_sheet(wb, report):
        """Create advertising sheet."""
        if not report.advertising_section:
            return

        ws = wb.create_sheet("Advertising")
        ad = report.advertising_section

        row = 1
        data = [
            ("Total Ad Spend (INR)", ad.total_ad_spend_inr),
            ("Attributed Sales (INR)", ad.total_attributed_sales_inr),
            ("Overall ROAS", f"{ad.overall_roas:.2f}x"),
            ("Overall ACOS (%)", f"{ad.overall_acos_pct:.1f}%"),
            ("Impressions", ad.impressions),
            ("Clicks", ad.clicks),
            ("CTR (%)", f"{ad.ctr_pct:.2f}%"),
            ("Attributed Units", ad.attributed_units),
            ("Attributed Orders", ad.attributed_orders),
        ]

        for field, value in data:
            ws[f"A{row}"] = field
            ws[f"B{row}"] = value
            row += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    @staticmethod
    def _create_profitability_sheet(wb, report):
        """Create profitability sheet."""
        if not report.profitability_section:
            return

        ws = wb.create_sheet("Profitability")
        prof = report.profitability_section

        row = 1
        ws[f"A{row}"] = "Profitability Summary"
        for key, val in ExcelRenderer._get_title_style().items():
            setattr(ws[f"A{row}"], key, val)

        row += 2
        data = [
            ("Total Net Sales (INR)", prof.total_net_sales_inr),
            ("Total Cost (INR)", prof.total_cost_inr),
            ("Total Contribution (INR)", prof.total_contribution_inr),
            ("Overall Profit Margin (%)", f"{prof.overall_profit_margin_pct:.1f}%"),
            ("Products (Healthy)", prof.products_healthy),
            ("Products (At Risk)", prof.products_at_risk),
            ("Products (Unprofitable)", prof.products_unprofitable),
        ]

        for field, value in data:
            ws[f"A{row}"] = field
            ws[f"B{row}"] = value
            row += 1

        row += 2
        ws[f"A{row}"] = "Cost Breakdown"
        for key, val in ExcelRenderer._get_title_style().items():
            setattr(ws[f"A{row}"], key, val)

        row += 1
        for cost_type, amount in prof.cost_breakdown.items():
            ws[f"A{row}"] = cost_type.replace("_", " ").title()
            ws[f"B{row}"] = amount
            row += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    @staticmethod
    def _create_quality_sheet(wb, report):
        """Create quality sheet."""
        if not report.quality_section:
            return

        ws = wb.create_sheet("Quality")
        quality = report.quality_section

        row = 1
        data = [
            ("Total Units Sold", quality.total_units_sold),
            ("Total Units Returned", quality.total_units_returned),
            ("Return Rate (%)", f"{quality.overall_return_rate_pct:.1f}%"),
            ("Total Refunds (INR)", quality.total_refund_amount_inr),
            ("Total Units Cancelled", quality.total_units_cancelled),
            ("Cancellation Rate (%)", f"{quality.overall_cancellation_rate_pct:.1f}%"),
        ]

        for field, value in data:
            ws[f"A{row}"] = field
            ws[f"B{row}"] = value
            row += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    @staticmethod
    def _create_insights_sheet(wb, report):
        """Create insights sheet."""
        if not report.insights:
            return

        ws = wb.create_sheet("Insights")

        row = 1
        headers = ["Title", "Description", "Priority", "Category"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            for key, val in ExcelRenderer._get_header_style().items():
                setattr(cell, key, val)

        row += 1
        for insight in report.insights:
            ws[f"A{row}"] = insight.title
            ws[f"B{row}"] = insight.description
            ws[f"C{row}"] = insight.priority
            ws[f"D{row}"] = insight.category
            ws[f"B{row}"].alignment = Alignment(wrap_text=True)
            row += 1

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 15

    @staticmethod
    def _create_recommendations_sheet(wb, report):
        """Create recommendations sheet."""
        if not report.recommendations:
            return

        ws = wb.create_sheet("Recommendations")

        row = 1
        headers = ["Action", "Rationale", "Owner", "Priority", "Timeline"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            for key, val in ExcelRenderer._get_header_style().items():
                setattr(cell, key, val)

        row += 1
        for rec in report.recommendations:
            ws[f"A{row}"] = rec.action
            ws[f"B{row}"] = rec.rationale
            ws[f"C{row}"] = rec.owner
            ws[f"D{row}"] = rec.priority
            ws[f"E{row}"] = rec.timeline or ""
            ws[f"A{row}"].alignment = Alignment(wrap_text=True)
            ws[f"B{row}"].alignment = Alignment(wrap_text=True)
            row += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 20

    @staticmethod
    def _create_metrics_sheet(wb, report):
        """Create overall metrics sheet."""
        ws = wb.create_sheet("Metrics")
        metrics = report.overall_metrics

        row = 1
        data = [
            ("Report Date", metrics.report_date.isoformat()),
            ("Total Orders", metrics.total_orders),
            ("Total Units Sold", metrics.total_units_sold),
            ("Total Net Sales (INR)", metrics.total_net_sales_inr),
            ("Total Gross Sales (INR)", metrics.total_gross_sales_inr),
            ("Total Ad Spend (INR)", metrics.total_ad_spend_inr),
            ("Total Organic Sales (INR)", metrics.total_organic_sales_inr),
            ("Organic Share (%)", f"{metrics.organic_share_pct:.1f}%"),
            ("Total Cost (INR)", metrics.total_cost_inr),
            ("Total Contribution (INR)", metrics.total_contribution_inr),
            ("Overall Profit Margin (%)", f"{metrics.overall_profit_margin_pct:.1f}%"),
            ("Total Return Rate (%)", f"{metrics.total_return_rate_pct:.1f}%"),
            ("Total Cancellation Rate (%)", f"{metrics.total_cancellation_rate_pct:.1f}%"),
            ("Product Count", metrics.product_count),
            ("Platform Count", metrics.platform_count),
        ]

        for field, value in data:
            ws[f"A{row}"] = field
            ws[f"B{row}"] = value
            row += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
