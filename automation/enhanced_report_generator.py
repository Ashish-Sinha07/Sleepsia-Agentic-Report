"""Enhanced Report Generator with PDF and Excel exports.

Generates business reports in multiple formats (JSON, PDF, Excel) from database data.
Integrates with email service for automated distribution.

Author: Claude Code
Date: 2026-08-24
"""

import logging
import os
from datetime import datetime, date
from typing import Dict, Optional, Tuple
from pathlib import Path

from sqlalchemy.orm import Session
from sqlalchemy import text

from backend.app.config import settings
from backend.app.services.report_service import ReportService
from automation.email_service import ReportEmailService

logger = logging.getLogger(__name__)


class EnhancedReportGenerator:
    """Generate business reports in multiple formats (JSON, PDF, Excel)."""

    def __init__(self, db: Session, output_dir: Optional[str] = None):
        """Initialize report generator.

        Args:
            db: Database session
            output_dir: Directory to save report files
        """
        self.db = db
        self.output_dir = Path(output_dir or "backend/reports")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.email_service = ReportEmailService()

        logger.info(f"Enhanced report generator initialized: {self.output_dir}")

    def generate_comprehensive_report(
        self,
        report_type: str = "executive_summary",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        include_pdf: bool = True,
        include_excel: bool = True,
    ) -> Dict[str, any]:
        """Generate comprehensive report in multiple formats.

        Args:
            report_type: Type of report to generate
            start_date: Start date for report period
            end_date: End date for report period
            include_pdf: Whether to generate PDF
            include_excel: Whether to generate Excel

        Returns:
            Dictionary with report metadata and file paths
        """
        try:
            start_date = start_date or date.today()
            end_date = end_date or date.today()

            logger.info(f"\n{'='*80}")
            logger.info(f"Generating comprehensive report: {report_type}")
            logger.info(f"Period: {start_date} to {end_date}")
            logger.info(f"{'='*80}\n")

            # Generate base JSON report
            logger.info("[1/3] Generating base report from database...")
            json_report = ReportService.generate_report(
                db=self.db,
                report_type=report_type,
                start_date=start_date,
                end_date=end_date,
            )
            report_id = json_report.get("report_id")
            logger.info(f"[OK] Report generated: {report_id}")

            report_files = {}
            report_files["json"] = f"{report_id}.json"

            # Generate PDF
            if include_pdf:
                logger.info("[2/3] Generating PDF report...")
                try:
                    pdf_bytes = self._generate_pdf(report_id, start_date, end_date)
                    pdf_filename = f"{report_id}.pdf"
                    pdf_path = self.output_dir / pdf_filename
                    pdf_path.write_bytes(pdf_bytes)
                    report_files["pdf"] = pdf_filename
                    logger.info(f"[OK] PDF generated: {pdf_filename} ({len(pdf_bytes)} bytes)")
                except Exception as e:
                    logger.warning(f"[WARN] PDF generation failed: {str(e)}")

            # Generate Excel
            if include_excel:
                logger.info("[3/3] Generating Excel report...")
                try:
                    excel_bytes = self._generate_excel(report_id, start_date, end_date)
                    excel_filename = f"{report_id}.xlsx"
                    excel_path = self.output_dir / excel_filename
                    excel_path.write_bytes(excel_bytes)
                    report_files["excel"] = excel_filename
                    logger.info(f"[OK] Excel generated: {excel_filename} ({len(excel_bytes)} bytes)")
                except Exception as e:
                    logger.warning(f"[WARN] Excel generation failed: {str(e)}")

            logger.info(f"\n{'='*80}")
            logger.info(f"Report generation complete: {len(report_files)} formats created")
            logger.info(f"Report ID: {report_id}")
            logger.info(f"{'='*80}\n")

            return {
                "success": True,
                "report_id": report_id,
                "report_type": report_type,
                "start_date": str(start_date),
                "end_date": str(end_date),
                "formats": report_files,
                "generated_at": datetime.now().isoformat(),
            }

        except Exception as e:
            logger.error(f"Report generation failed: {str(e)}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "generated_at": datetime.now().isoformat(),
            }

    def _generate_pdf(self, report_id: str, start_date: date, end_date: date) -> bytes:
        """Generate PDF report using reportlab.

        Args:
            report_id: Report ID
            start_date: Start date
            end_date: End date

        Returns:
            PDF bytes
        """
        return self._generate_pdf_with_reportlab(report_id, start_date, end_date)

    def _generate_pdf_with_reportlab(
        self, report_id: str, start_date: date, end_date: date
    ) -> bytes:
        """Generate PDF using reportlab as fallback."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            from io import BytesIO

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5 * inch)
            story = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=18,
                textColor=colors.HexColor("#2C3E50"),
                spaceAfter=12,
            )

            story.append(Paragraph(f"Sleepsia Business Report - {report_id}", title_style))
            story.append(Paragraph(f"Period: {start_date} to {end_date}", styles["Normal"]))
            story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
            story.append(Spacer(1, 0.3 * inch))

            # Get data from database
            logger.info("Fetching report data from database...")
            kpi_data = self._fetch_kpi_data(start_date, end_date)

            # Add KPI section
            story.append(Paragraph("Key Performance Indicators", styles["Heading2"]))
            if kpi_data:
                kpi_table_data = [
                    ["Metric", "Value"],
                    ["Revenue", f"₹{kpi_data.get('revenue', 0):,.0f}"],
                    ["Net Revenue", f"₹{kpi_data.get('net_revenue', 0):,.0f}"],
                    ["Profit", f"₹{kpi_data.get('profit', 0):,.0f}"],
                    ["Profit Margin", f"{kpi_data.get('profit_margin', 0):.1f}%"],
                    ["Orders", f"{kpi_data.get('orders', 0):,}"],
                    ["Units Sold", f"{kpi_data.get('units', 0):,}"],
                    ["ROAS", f"{kpi_data.get('roas', 0):.2f}x"],
                    ["Ad Spend", f"₹{kpi_data.get('ad_spend', 0):,.0f}"],
                ]

                kpi_table = Table(kpi_table_data, colWidths=[3 * inch, 2 * inch])
                kpi_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 11),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
                ]))
                story.append(kpi_table)

            story.append(Spacer(1, 0.3 * inch))

            # Get platform data
            logger.info("Fetching platform data...")
            platform_data = self._fetch_platform_data(start_date, end_date)

            if platform_data:
                story.append(PageBreak())
                story.append(Paragraph("Platform Performance", styles["Heading2"]))

                platform_table_data = [
                    ["Platform", "Revenue", "Profit", "Margin (%)", "Orders"]
                ]

                for platform in platform_data:
                    platform_table_data.append([
                        platform.get("platform", ""),
                        f"₹{platform.get('revenue', 0):,.0f}",
                        f"₹{platform.get('profit', 0):,.0f}",
                        f"{platform.get('margin', 0):.1f}%",
                        f"{platform.get('orders', 0):,}",
                    ])

                platform_table = Table(platform_table_data, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 1.2 * inch, 1 * inch])
                platform_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
                ]))
                story.append(platform_table)

            # Get product data
            logger.info("Fetching product data...")
            product_data = self._fetch_product_data(start_date, end_date)

            if product_data:
                story.append(PageBreak())
                story.append(Paragraph("Top Products by Profit", styles["Heading2"]))

                product_table_data = [
                    ["Product", "Revenue", "Profit", "Margin (%)", "Units"]
                ]

                for product in product_data[:15]:  # Top 15 products
                    product_table_data.append([
                        product.get("product", ""),
                        f"₹{product.get('revenue', 0):,.0f}",
                        f"₹{product.get('profit', 0):,.0f}",
                        f"{product.get('margin', 0):.1f}%",
                        f"{product.get('units', 0):,}",
                    ])

                product_table = Table(product_table_data, colWidths=[1.8 * inch, 1.3 * inch, 1.3 * inch, 1.2 * inch, 0.8 * inch])
                product_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
                ]))
                story.append(product_table)

            # Build PDF
            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue()

        except Exception as e:
            logger.error(f"PDF generation with reportlab failed: {str(e)}")
            raise

    def _generate_html_report(self, report_id: str, start_date: date, end_date: date) -> str:
        """Generate HTML report content."""
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 20px;
                    color: #333;
                }}
                h1 {{
                    color: #2C3E50;
                    border-bottom: 2px solid #34495E;
                    padding-bottom: 10px;
                }}
                h2 {{
                    color: #34495E;
                    margin-top: 30px;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 15px 0;
                }}
                th {{
                    background-color: #34495E;
                    color: white;
                    padding: 12px;
                    text-align: left;
                    font-weight: bold;
                }}
                td {{
                    border: 1px solid #ddd;
                    padding: 10px;
                }}
                tr:nth-child(even) {{
                    background-color: #f8f9fa;
                }}
                .metric-label {{
                    font-weight: bold;
                    width: 40%;
                }}
                .summary {{
                    background-color: #ecf0f1;
                    padding: 15px;
                    border-radius: 5px;
                    margin: 10px 0;
                }}
            </style>
        </head>
        <body>
            <h1>Sleepsia Business Report</h1>
            <p><strong>Report ID:</strong> {report_id}</p>
            <p><strong>Period:</strong> {start_date} to {end_date}</p>
            <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>

            <h2>Executive Summary</h2>
            <div class="summary">
                <p>This comprehensive business report provides detailed insights into
                platform performance, product profitability, advertising efficiency, and
                key business metrics for the Sleepsia organization.</p>
            </div>

            <h2>Key Performance Indicators</h2>
            <table>
                <tr>
                    <th>Metric</th>
                    <th>Value</th>
                </tr>
        """

        # Add KPI data
        kpi_data = self._fetch_kpi_data(start_date, end_date)
        if kpi_data:
            html += f"""
                <tr>
                    <td class="metric-label">Revenue</td>
                    <td>₹{kpi_data.get('revenue', 0):,.0f}</td>
                </tr>
                <tr>
                    <td class="metric-label">Profit</td>
                    <td>₹{kpi_data.get('profit', 0):,.0f}</td>
                </tr>
                <tr>
                    <td class="metric-label">Profit Margin</td>
                    <td>{kpi_data.get('profit_margin', 0):.1f}%</td>
                </tr>
                <tr>
                    <td class="metric-label">Orders</td>
                    <td>{kpi_data.get('orders', 0):,}</td>
                </tr>
                <tr>
                    <td class="metric-label">Units Sold</td>
                    <td>{kpi_data.get('units', 0):,}</td>
                </tr>
                <tr>
                    <td class="metric-label">ROAS</td>
                    <td>{kpi_data.get('roas', 0):.2f}x</td>
                </tr>
                <tr>
                    <td class="metric-label">Ad Spend</td>
                    <td>₹{kpi_data.get('ad_spend', 0):,.0f}</td>
                </tr>
            """

        html += """
            </table>
        </body>
        </html>
        """
        return html

    def _generate_excel(self, report_id: str, start_date: date, end_date: date) -> bytes:
        """Generate Excel report."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from io import BytesIO

            wb = Workbook()
            wb.remove(wb.active)

            # Summary sheet
            ws_summary = wb.create_sheet("Summary", 0)
            ws_summary["A1"] = "Sleepsia Business Report"
            ws_summary["A1"].font = Font(bold=True, size=14, color="2C3E50")

            ws_summary["A3"] = "Report Information"
            ws_summary["A3"].font = Font(bold=True, size=11)

            row = 4
            summary_data = [
                ("Report ID", report_id),
                ("Report Type", "Executive Summary"),
                ("Period", f"{start_date} to {end_date}"),
                ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ]

            for label, value in summary_data:
                ws_summary[f"A{row}"] = label
                ws_summary[f"B{row}"] = value
                ws_summary[f"A{row}"].font = Font(bold=True)
                row += 1

            # KPI sheet
            ws_kpi = wb.create_sheet("KPIs", 1)
            ws_kpi["A1"] = "Key Performance Indicators"
            ws_kpi["A1"].font = Font(bold=True, size=12, color="FFFFFF")
            ws_kpi["A1"].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")

            row = 3
            kpi_data = self._fetch_kpi_data(start_date, end_date)
            if kpi_data:
                kpi_rows = [
                    ("Revenue", f"₹{kpi_data.get('revenue', 0):,.0f}"),
                    ("Net Revenue", f"₹{kpi_data.get('net_revenue', 0):,.0f}"),
                    ("Profit", f"₹{kpi_data.get('profit', 0):,.0f}"),
                    ("Profit Margin", f"{kpi_data.get('profit_margin', 0):.1f}%"),
                    ("Orders", f"{kpi_data.get('orders', 0):,}"),
                    ("Units Sold", f"{kpi_data.get('units', 0):,}"),
                    ("ROAS", f"{kpi_data.get('roas', 0):.2f}x"),
                    ("Ad Spend", f"₹{kpi_data.get('ad_spend', 0):,.0f}"),
                    ("ACOS", f"{kpi_data.get('acos', 0):.1f}%"),
                ]

                for label, value in kpi_rows:
                    ws_kpi[f"A{row}"] = label
                    ws_kpi[f"B{row}"] = value
                    ws_kpi[f"A{row}"].font = Font(bold=True)
                    row += 1

            ws_kpi.column_dimensions["A"].width = 20
            ws_kpi.column_dimensions["B"].width = 20

            # Platforms sheet
            ws_platforms = wb.create_sheet("Platforms", 2)
            ws_platforms["A1"] = "Platform Performance"
            ws_platforms["A1"].font = Font(bold=True, size=12, color="FFFFFF")
            ws_platforms["A1"].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")

            headers = ["Platform", "Revenue", "Profit", "Margin (%)", "Orders"]
            for col, header in enumerate(headers, 1):
                cell = ws_platforms.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")

            row = 4
            platform_data = self._fetch_platform_data(start_date, end_date)
            if platform_data:
                for platform in platform_data:
                    ws_platforms[f"A{row}"] = platform.get("platform", "")
                    ws_platforms[f"B{row}"] = platform.get("revenue", 0)
                    ws_platforms[f"C{row}"] = platform.get("profit", 0)
                    ws_platforms[f"D{row}"] = platform.get("margin", 0)
                    ws_platforms[f"E{row}"] = platform.get("orders", 0)
                    row += 1

            for col in range(1, 6):
                ws_platforms.column_dimensions[chr(64 + col)].width = 15

            # Products sheet
            ws_products = wb.create_sheet("Products", 3)
            ws_products["A1"] = "Top Products by Profit"
            ws_products["A1"].font = Font(bold=True, size=12, color="FFFFFF")
            ws_products["A1"].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")

            headers = ["Product", "Revenue", "Profit", "Margin (%)", "Units"]
            for col, header in enumerate(headers, 1):
                cell = ws_products.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")

            row = 4
            product_data = self._fetch_product_data(start_date, end_date)
            if product_data:
                for product in product_data[:20]:  # Top 20 products
                    ws_products[f"A{row}"] = product.get("product", "")
                    ws_products[f"B{row}"] = product.get("revenue", 0)
                    ws_products[f"C{row}"] = product.get("profit", 0)
                    ws_products[f"D{row}"] = product.get("margin", 0)
                    ws_products[f"E{row}"] = product.get("units", 0)
                    row += 1

            for col in range(1, 6):
                ws_products.column_dimensions[chr(64 + col)].width = 15

            # Save to bytes
            stream = BytesIO()
            wb.save(stream)
            stream.seek(0)
            return stream.getvalue()

        except Exception as e:
            logger.error(f"Excel generation failed: {str(e)}")
            raise

    def _fetch_kpi_data(self, start_date: date, end_date: date) -> Dict:
        """Fetch KPI data from database."""
        try:
            query = """
            SELECT
                SUM(total_gross_sales) as revenue,
                SUM(total_net_sales) as net_revenue,
                SUM(total_contribution) as profit,
                AVG(overall_profit_margin_pct) as profit_margin,
                SUM(total_orders) as orders,
                SUM(total_units_sold) as units,
                AVG(overall_roas) as roas,
                SUM(total_ad_spend) as ad_spend,
                AVG(overall_acos_pct) as acos
            FROM vw_daily_kpi_summary
            WHERE date BETWEEN :start_date AND :end_date
            """

            result = self.db.execute(
                text(query),
                {"start_date": start_date, "end_date": end_date}
            ).fetchone()

            if result:
                return {
                    "revenue": float(result[0]) if result[0] else 0,
                    "net_revenue": float(result[1]) if result[1] else 0,
                    "profit": float(result[2]) if result[2] else 0,
                    "profit_margin": float(result[3]) if result[3] else 0,
                    "orders": int(result[4]) if result[4] else 0,
                    "units": int(result[5]) if result[5] else 0,
                    "roas": float(result[6]) if result[6] else 0,
                    "ad_spend": float(result[7]) if result[7] else 0,
                    "acos": float(result[8]) if result[8] else 0,
                }
            return {}

        except Exception as e:
            logger.warning(f"Failed to fetch KPI data: {str(e)}")
            return {}

    def _fetch_platform_data(self, start_date: date, end_date: date) -> list:
        """Fetch platform performance data."""
        try:
            query = """
            SELECT
                platform_name,
                SUM(gross_sales) as revenue,
                SUM(contribution) as profit,
                AVG(profit_margin_pct) as margin,
                SUM(orders) as orders
            FROM vw_platform_performance
            WHERE date BETWEEN :start_date AND :end_date
            GROUP BY platform_name
            ORDER BY revenue DESC
            """

            results = self.db.execute(
                text(query),
                {"start_date": start_date, "end_date": end_date}
            ).fetchall()

            return [
                {
                    "platform": row[0],
                    "revenue": float(row[1]) if row[1] else 0,
                    "profit": float(row[2]) if row[2] else 0,
                    "margin": float(row[3]) if row[3] else 0,
                    "orders": int(row[4]) if row[4] else 0,
                }
                for row in results
            ]

        except Exception as e:
            logger.warning(f"Failed to fetch platform data: {str(e)}")
            return []

    def _fetch_product_data(self, start_date: date, end_date: date) -> list:
        """Fetch product performance data."""
        try:
            query = """
            SELECT
                product_name,
                SUM(gross_sales) as revenue,
                SUM(contribution) as profit,
                AVG(profit_margin_pct) as margin,
                SUM(units_sold) as units
            FROM vw_product_performance
            WHERE date BETWEEN :start_date AND :end_date
            GROUP BY product_name
            ORDER BY profit DESC
            LIMIT 30
            """

            results = self.db.execute(
                text(query),
                {"start_date": start_date, "end_date": end_date}
            ).fetchall()

            return [
                {
                    "product": row[0],
                    "revenue": float(row[1]) if row[1] else 0,
                    "profit": float(row[2]) if row[2] else 0,
                    "margin": float(row[3]) if row[3] else 0,
                    "units": int(row[4]) if row[4] else 0,
                }
                for row in results
            ]

        except Exception as e:
            logger.warning(f"Failed to fetch product data: {str(e)}")
            return []

    def send_report_via_email(
        self,
        report_id: str,
        recipients: list,
        cc: Optional[list] = None,
        bcc: Optional[list] = None,
    ) -> bool:
        """Send generated report files via email.

        Args:
            report_id: Report ID
            recipients: List of recipient emails
            cc: Optional list of CC recipients
            bcc: Optional list of BCC recipients

        Returns:
            True if email sent successfully
        """
        try:
            logger.info(f"\nSending report {report_id} via email...")

            # Find report files
            attachments = {}

            for ext in ["json", "pdf", "xlsx"]:
                file_path = self.output_dir / f"{report_id}.{ext}"
                if file_path.exists():
                    with open(file_path, "rb") as f:
                        filename = file_path.name
                        attachments[filename] = f.read()
                    logger.info(f"[OK] Attached: {filename}")

            if not attachments:
                logger.error("No report files found to attach")
                return False

            # Send via email service
            success = self.email_service.send_report(
                subject=f"Sleepsia Business Report - {report_id}",
                body=f"""
Dear Recipient,

Please find attached your comprehensive business report for Sleepsia.

Report Includes:
• Executive Summary with Key Performance Indicators
• Platform Performance Analysis
• Product Profitability Analysis
• Advertising ROI and Efficiency Metrics
• Detailed Data Sheets (Excel)

Report ID: {report_id}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Formats Included:
- JSON (Raw Data)
- PDF (Formatted Report)
- Excel (Data Sheets)

If you have any questions or need additional analysis, please reach out.

Best regards,
Sleepsia Analytics System
                """.strip(),
                recipients=recipients,
                cc=cc,
                bcc=bcc,
                attachments=attachments,
            )

            if success:
                logger.info(f"[OK] Report sent to {len(recipients)} recipient(s)")
            else:
                logger.error("Failed to send report via email")

            return success

        except Exception as e:
            logger.error(f"Email sending failed: {str(e)}", exc_info=True)
            return False
