"""Excel/CSV/Markdown ingestion pipeline for the RAG knowledge base.

Only files/sheets representing static business knowledge (policies,
guidelines, config thresholds, SOPs) are meant to flow through here.
Transactional/operational data keeps flowing through the existing ETL
pipeline (backend/etl/loader.py) into MySQL - this module never touches
that path and never writes to MySQL.

Uploaded file content is untrusted input: validated for type/size, filenames
are sanitized against path traversal, and nothing here ever executes,
evaluates, or shells out based on file content - it is only ever parsed into
plain text chunks for the vector store.
"""

import csv as csv_module
import io
import logging
import os
import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional

import openpyxl

from app.rag.chunking import Chunk, chunk_markdown, chunk_sheet_rows
from app.rag.vector_store import VectorStore, make_chunk_id

logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv", ".md"}
MAX_SHEETS_PER_WORKBOOK = 50
MAX_ROWS_PER_SHEET = 5000


class IngestionError(ValueError):
    """Raised for invalid uploads: bad extension, oversized, empty, or malformed content."""


@dataclass
class IngestionSummary:
    filename: str
    sheets_processed: int
    documents_created: int
    chunks_created: int
    status: str
    error: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "filename": self.filename,
            "sheets_processed": self.sheets_processed,
            "documents_created": self.documents_created,
            "chunks_created": self.chunks_created,
            "status": self.status,
            "error": self.error,
        }


def sanitize_filename(filename: str) -> str:
    """Strip any directory components and disallow path traversal."""
    base = os.path.basename((filename or "").strip())
    base = base.replace("..", "")
    base = re.sub(r"[^A-Za-z0-9_.\- ]", "_", base)
    return base[:150] or "upload"


def validate_upload(filename: str, content: bytes, max_upload_mb: int) -> str:
    """Validate extension + size. Returns the sanitized filename or raises IngestionError."""
    safe_name = sanitize_filename(filename)
    ext = os.path.splitext(safe_name)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise IngestionError(
            f"Unsupported file type '{ext}'. Allowed types: {', '.join(sorted(ALLOWED_EXTENSIONS))}"
        )
    if len(content) == 0:
        raise IngestionError("File is empty")
    max_bytes = max_upload_mb * 1024 * 1024
    if len(content) > max_bytes:
        raise IngestionError(f"File exceeds the {max_upload_mb}MB upload limit")
    return safe_name


class IngestionService:
    """Parses, chunks, embeds, and stores knowledge documents."""

    def __init__(self, vector_store: VectorStore):
        self._store = vector_store

    def ingest_bytes(
        self,
        filename: str,
        content: bytes,
        max_upload_mb: int = 10,
        knowledge_sheets: Optional[List[str]] = None,
    ) -> IngestionSummary:
        safe_name = validate_upload(filename, content, max_upload_mb)
        ext = os.path.splitext(safe_name)[1].lower()
        try:
            if ext == ".md":
                return self._ingest_markdown_bytes(safe_name, content)
            if ext in (".xlsx", ".xls"):
                return self._ingest_workbook_bytes(safe_name, content, knowledge_sheets)
            if ext == ".csv":
                return self._ingest_csv_bytes(safe_name, content)
        except IngestionError:
            raise
        except Exception as e:
            logger.exception("Ingestion failed for %s", safe_name)
            return IngestionSummary(safe_name, 0, 0, 0, "error", error=str(e))
        raise IngestionError(f"Unsupported file type '{ext}'")

    def _store_chunks(self, source_file: str, chunks: List[Chunk]) -> int:
        # Re-ingesting the same source file replaces its previous chunks
        # rather than accumulating duplicates alongside them.
        self._store.delete_documents(where={"source_file": source_file})
        if not chunks:
            return 0
        ids, texts, metadatas = [], [], []
        for i, chunk in enumerate(chunks):
            section = str(chunk.metadata.get("sheet_name") or chunk.metadata.get("section") or "chunk")
            ids.append(make_chunk_id(source_file, section, i))
            texts.append(chunk.text)
            metadatas.append(chunk.metadata)
        self._store.add_documents(ids=ids, texts=texts, metadatas=metadatas)
        return len(chunks)

    def _ingest_markdown_bytes(self, safe_name: str, content: bytes) -> IngestionSummary:
        text = content.decode("utf-8", errors="replace")
        chunks = chunk_markdown(text, source_file=safe_name)
        count = self._store_chunks(safe_name, chunks)
        return IngestionSummary(
            safe_name, sheets_processed=1, documents_created=1, chunks_created=count, status="success"
        )

    def _ingest_csv_bytes(self, safe_name: str, content: bytes) -> IngestionSummary:
        text = content.decode("utf-8", errors="replace")
        reader = csv_module.DictReader(io.StringIO(text))
        rows = []
        for row in reader:
            rows.append(dict(row))
            if len(rows) >= MAX_ROWS_PER_SHEET:
                break
        chunks = chunk_sheet_rows(rows, sheet_name=os.path.splitext(safe_name)[0], source_file=safe_name)
        count = self._store_chunks(safe_name, chunks)
        return IngestionSummary(
            safe_name,
            sheets_processed=1,
            documents_created=len(rows),
            chunks_created=count,
            status="success",
        )

    def _ingest_workbook_bytes(
        self, safe_name: str, content: bytes, knowledge_sheets: Optional[List[str]] = None
    ) -> IngestionSummary:
        """Ingest workbook sheets as knowledge documents.

        `knowledge_sheets`, when given, restricts ingestion to that explicit
        allowlist (used for the bundled config/policy sheets in the main
        Sleepsia workbook, which also contains transactional ETL sheets we
        must never index as "knowledge"). Ad hoc admin uploads with no
        allowlist ingest every sheet in the file - an admin uploading a
        document through the knowledge endpoint is presumed to be curating
        reference/policy content, not raw transactional data.
        """
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        try:
            sheet_names = wb.sheetnames[:MAX_SHEETS_PER_WORKBOOK]
            if knowledge_sheets is not None:
                sheet_names = [s for s in sheet_names if s in knowledge_sheets]

            all_chunks: List[Chunk] = []
            documents_created = 0
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)
                try:
                    header_row = next(rows_iter)
                except StopIteration:
                    continue
                header = [
                    (str(h).strip() if h is not None and str(h).strip() else f"col_{i}")
                    for i, h in enumerate(header_row)
                ]
                row_dicts: List[Dict[str, Any]] = []
                for row in rows_iter:
                    if all(cell is None or str(cell).strip() == "" for cell in row):
                        continue
                    row_dicts.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
                    if len(row_dicts) >= MAX_ROWS_PER_SHEET:
                        break
                documents_created += len(row_dicts)
                all_chunks.extend(
                    chunk_sheet_rows(row_dicts, sheet_name=sheet_name, source_file=safe_name)
                )
        finally:
            wb.close()

        count = self._store_chunks(safe_name, all_chunks)
        return IngestionSummary(
            safe_name,
            sheets_processed=len(sheet_names),
            documents_created=documents_created,
            chunks_created=count,
            status="success",
        )

    def delete_source(self, source_file: str) -> int:
        return self._store.delete_documents(where={"source_file": sanitize_filename(source_file)})

    def list_sources(self) -> List[Dict[str, Any]]:
        return self._store.list_sources()
