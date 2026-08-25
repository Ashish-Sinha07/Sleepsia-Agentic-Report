"""Service for generating business reports.

Produces a "Sleepsia Omni-Channel Audit" style report (channel performance,
product performance by platform, inventory & warehouse operations,
consolidated SKU rollups, a consolidated P&L, and agent-generated audit
notes) rendered to JSON, PDF, or Excel from real data in the `sleepsia`
database. Every figure comes from the same views/tables the rest of the
backend (kpi_service, platform_service, product_service, ...) already uses;
nothing is invented, and any template column with no real data source in
this schema is simply omitted rather than faked.
"""

import os
import re
import uuid
import logging
from collections import OrderedDict, defaultdict
from datetime import datetime, date, timedelta
from typing import Optional, Dict, List, Any, Tuple
import json

from sqlalchemy import text
from sqlalchemy.orm import Session

from analytics.models import ProductMetrics, PlatformMetrics
from agents.analysis_agent import DataAnalysisAgent
from agents.insight_recommendation_agent import InsightRecommendationAgent

logger = logging.getLogger(__name__)


def _f(value: Any, default: float = 0.0) -> float:
    """Safely coerce a DB value (Decimal/None/str) to float."""
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _i(value: Any, default: int = 0) -> int:
    """Safely coerce a DB value to int."""
    if value is None:
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _pct(numerator: float, denominator: float) -> float:
    """Percentage helper that never divides by zero."""
    if not denominator:
        return 0.0
    return (numerator / denominator) * 100.0


def _ratio(numerator: float, denominator: float) -> float:
    """Ratio helper (e.g. ROAS) that never divides by zero."""
    if not denominator:
        return 0.0
    return numerator / denominator


class ReportService:
    """Generate and manage business reports."""

    REPORT_TYPES = {
        "executive_summary": "Executive Summary Report",
        "platform_analysis": "Platform Performance Analysis",
        "product_analysis": "Product Profitability Analysis",
        "profitability": "Detailed Profitability Analysis",
        "advertising": "Advertising ROI Analysis",
        "inventory": "Warehouse & Inventory Analysis",
        "management_monthly": "Management Monthly Report",
    }

    REPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "reports")

    # ------------------------------------------------------------------
    # Public entry points
    # ------------------------------------------------------------------

    @staticmethod
    def _ensure_reports_dir():
        """Ensure reports directory exists."""
        os.makedirs(ReportService.REPORTS_DIR, exist_ok=True)

    @staticmethod
    def generate_report(
        db: Session,
        report_type: str,
        start_date: date,
        end_date: date,
        format: str = "pdf",
        include_recommendations: bool = True,
        platform_filter: Optional[str] = None,
        warehouse_filter: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Generate a comprehensive business report."""

        if report_type not in ReportService.REPORT_TYPES:
            raise ValueError(f"Unknown report type: {report_type}")

        ReportService._ensure_reports_dir()

        report_id = f"REP-{uuid.uuid4().hex[:8].upper()}"
        report_title = ReportService.REPORT_TYPES[report_type]

        # Collect data based on report type
        report_data = ReportService._collect_report_data(
            db=db,
            report_type=report_type,
            start_date=start_date,
            end_date=end_date,
            platform_filter=platform_filter,
            warehouse_filter=warehouse_filter,
        )

        # Generate report content
        report_content = ReportService._generate_report_content(
            report_id=report_id,
            report_type=report_type,
            report_title=report_title,
            data=report_data,
            start_date=start_date,
            end_date=end_date,
            include_recommendations=include_recommendations,
        )

        requested_format = format.lower()
        if requested_format == "excel":
            requested_format = "xlsx"
        if requested_format not in {"json", "pdf", "xlsx"}:
            raise ValueError("Format must be json, pdf, or excel")

        # JSON is the canonical report; other formats are rendered from it.
        json_path = os.path.join(ReportService.REPORTS_DIR, f"{report_id}.json")
        ReportService._ensure_reports_dir()
        with open(json_path, 'w') as f:
            json.dump(report_content, f, indent=2, default=str)

        file_path = json_path
        if requested_format == "pdf":
            file_path = os.path.join(ReportService.REPORTS_DIR, f"{report_id}.pdf")
            ReportService._render_pdf(report_content, file_path)
        elif requested_format == "xlsx":
            file_path = os.path.join(ReportService.REPORTS_DIR, f"{report_id}.xlsx")
            ReportService._render_excel(report_content, file_path)

        return {
            "report_id": report_id,
            "report_type": report_type,
            "created_at": datetime.now().isoformat(),
            "start_date": start_date,
            "end_date": end_date,
            "status": "completed",
            "file_size": os.path.getsize(file_path),
            "format": requested_format,
            "download_url": f"/api/reports/{report_id}/download?format={requested_format}"
        }

    # ------------------------------------------------------------------
    # Data collection
    # ------------------------------------------------------------------

    @staticmethod
    def _collect_report_data(
        db: Session,
        report_type: str,
        start_date: date,
        end_date: date,
        platform_filter: Optional[str] = None,
        warehouse_filter: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Collect data for report based on type."""

        data = {
            "period": {"start_date": start_date.isoformat(), "end_date": end_date.isoformat()},
            "generated_at": datetime.now().isoformat(),
        }

        data["kpi_summary"] = ReportService._get_kpi_summary(db, start_date, end_date, platform_filter)

        if report_type in ("executive_summary", "management_monthly"):
            # Full omni-channel audit: every section.
            data["channel_data"] = ReportService._get_channel_data(db, start_date, end_date, platform_filter)
            data["product_by_platform"] = ReportService._get_product_by_platform_data(
                db, start_date, end_date, platform_filter
            )
            sku_sales = ReportService._get_sku_sales_summary(db, start_date, end_date, platform_filter)
            warehouse_rows = ReportService._get_warehouse_ops_rows(db, start_date, end_date, warehouse_filter)
            data["warehouse_operations"] = warehouse_rows
            data["inventory_position"] = ReportService._get_inventory_position_data(sku_sales, warehouse_rows)
            data["consolidated_sku"] = ReportService._get_consolidated_sku_data(sku_sales, warehouse_rows)
            data["pnl"] = ReportService._get_pnl_data(db, start_date, end_date, platform_filter)
            audit_notes, audit_source = ReportService._get_audit_notes(
                db, start_date, end_date, platform_filter
            )
            data["audit_notes"] = audit_notes
            data["audit_notes_source"] = audit_source

        elif report_type == "platform_analysis":
            data["channel_data"] = ReportService._get_channel_data(db, start_date, end_date, platform_filter)
            data["product_by_platform"] = ReportService._get_product_by_platform_data(
                db, start_date, end_date, platform_filter
            )

        elif report_type == "product_analysis":
            sku_sales = ReportService._get_sku_sales_summary(db, start_date, end_date, platform_filter)
            warehouse_rows = ReportService._get_warehouse_ops_rows(db, start_date, end_date, warehouse_filter)
            data["consolidated_sku"] = ReportService._get_consolidated_sku_data(sku_sales, warehouse_rows)
            data["product_by_platform"] = ReportService._get_product_by_platform_data(
                db, start_date, end_date, platform_filter
            )

        elif report_type == "profitability":
            data["pnl"] = ReportService._get_pnl_data(db, start_date, end_date, platform_filter)
            data["profitability_data"] = ReportService._get_profitability_trend(db, start_date, end_date)

        elif report_type == "advertising":
            data["advertising_data"] = ReportService._get_advertising_data(db, start_date, end_date)

        elif report_type == "inventory":
            sku_sales = ReportService._get_sku_sales_summary(db, start_date, end_date, platform_filter)
            warehouse_rows = ReportService._get_warehouse_ops_rows(db, start_date, end_date, warehouse_filter)
            data["warehouse_operations"] = warehouse_rows
            data["inventory_position"] = ReportService._get_inventory_position_data(sku_sales, warehouse_rows)

        return data

    @staticmethod
    def _get_kpi_summary(
        db: Session,
        start_date: date,
        end_date: date,
        platform_filter: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Aggregate KPI summary from vw_product_platform_daily.

        Mirrors KpiService.get_daily_kpis (deriving ratios from summed
        totals rather than averaging pre-computed daily percentages) so the
        report's headline numbers match the dashboard's KPI page exactly.
        Uses vw_product_platform_daily (which carries platform_id) instead
        of the platform-less vw_daily_kpi_summary so an optional platform
        filter can actually be applied.
        """
        query = """
        SELECT
            COALESCE(SUM(orders), 0) as total_orders,
            COALESCE(SUM(units_sold), 0) as total_units_sold,
            COALESCE(SUM(gross_sales), 0) as total_gross_sales,
            COALESCE(SUM(net_sales), 0) as total_net_sales,
            COALESCE(SUM(ad_spend), 0) as total_ad_spend,
            COALESCE(SUM(ad_attributed_sales), 0) as total_ad_sales,
            COALESCE(SUM(contribution_inr), 0) as total_contribution
        FROM vw_product_platform_daily
        WHERE date BETWEEN :start_date AND :end_date
        """
        params: Dict[str, Any] = {"start_date": start_date, "end_date": end_date}
        if platform_filter:
            query += " AND platform_id = :platform_filter"
            params["platform_filter"] = platform_filter

        try:
            result = db.execute(text(query), params).fetchone()
        except Exception as e:
            logger.error(f"KPI query failed: {str(e)}", exc_info=True)
            result = None

        if not result:
            return {k: 0 for k in [
                "revenue", "net_revenue", "profit", "profit_margin_pct",
                "orders", "units_sold", "roas", "ad_spend",
            ]}

        net_sales = _f(result[3])
        ad_sales = _f(result[5])
        ad_spend = _f(result[4])
        contribution = _f(result[6])

        return {
            "revenue": _f(result[2]),
            "net_revenue": net_sales,
            "profit": contribution,
            "profit_margin_pct": _pct(contribution, net_sales),
            "orders": _i(result[0]),
            "units_sold": _i(result[1]),
            "roas": _ratio(ad_sales, ad_spend),
            "ad_spend": ad_spend,
        }

    # -- Section 1: Channel Performance -------------------------------

    @staticmethod
    def _get_channel_data(
        db: Session,
        start_date: date,
        end_date: date,
        platform_filter: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        """One row per platform that actually has data in the period.

        Columns match the audit template minus "OTIF %" (there is no
        on-time-in-full/delivery table anywhere in this schema, so it is
        dropped rather than faked).
        """
        query = """
        SELECT
            platform_id,
            platform,
            COALESCE(SUM(orders), 0) AS orders,
            COALESCE(SUM(units_sold), 0) AS units,
            COALESCE(SUM(gross_sales), 0) AS gross_sales,
            COALESCE(SUM(refund_amount), 0) AS returns_amount,
            COALESCE(SUM(net_sales), 0) AS net_sales,
            COALESCE(SUM(ad_spend), 0) AS ad_spend,
            COALESCE(SUM(contribution_inr), 0) AS profit
        FROM vw_product_platform_daily
        WHERE date BETWEEN :start_date AND :end_date
        """
        params: Dict[str, Any] = {"start_date": start_date, "end_date": end_date}
        if platform_filter:
            query += " AND platform_id = :platform_filter"
            params["platform_filter"] = platform_filter
        query += " GROUP BY platform_id, platform ORDER BY gross_sales DESC"

        try:
            results = db.execute(text(query), params).fetchall()
        except Exception as e:
            logger.error(f"Channel data query failed: {str(e)}", exc_info=True)
            return []

        rows = []
        for row in results:
            gross_sales = _f(row[4])
            returns_amount = _f(row[5])
            net_sales = _f(row[6])
            ad_spend = _f(row[7])
            profit = _f(row[8])
            net_revenue = net_sales - returns_amount

            rows.append({
                "platform_id": row[0],
                "platform": row[1],
                "orders": _i(row[2]),
                "units": _i(row[3]),
                "gross_sales": gross_sales,
                "returns_amount": returns_amount,
                "net_revenue": net_revenue,
                "ad_spend": ad_spend,
                "tacos_pct": _pct(ad_spend, gross_sales),
                "profit": profit,
                "margin_pct": _pct(profit, net_sales),
            })
        return rows

    # -- Section 2: Product Performance by Platform --------------------

    @staticmethod
    def _get_product_by_platform_data(
        db: Session,
        start_date: date,
        end_date: date,
        platform_filter: Optional[str] = None,
    ) -> "OrderedDict[str, List[Dict[str, Any]]]":
        """One product table per platform.

        Organic units = units_sold - ad_attributed_units;
        Paid units = ad_attributed_units (per the ETL's attribution rule).
        """
        query = """
        SELECT
            platform_id,
            platform,
            sku,
            product_name,
            COALESCE(SUM(units_sold), 0) AS units,
            COALESCE(SUM(gross_sales), 0) AS gross_sales,
            COALESCE(SUM(net_sales), 0) AS net_sales,
            COALESCE(SUM(units_returned), 0) AS units_returned,
            COALESCE(SUM(attributed_units), 0) AS paid_units,
            COALESCE(SUM(ad_spend), 0) AS ad_spend,
            COALESCE(SUM(contribution_inr), 0) AS profit
        FROM vw_product_platform_daily
        WHERE date BETWEEN :start_date AND :end_date
        """
        params: Dict[str, Any] = {"start_date": start_date, "end_date": end_date}
        if platform_filter:
            query += " AND platform_id = :platform_filter"
            params["platform_filter"] = platform_filter
        query += " GROUP BY platform_id, platform, sku, product_name ORDER BY platform, profit DESC"

        try:
            results = db.execute(text(query), params).fetchall()
        except Exception as e:
            logger.error(f"Product-by-platform query failed: {str(e)}", exc_info=True)
            return OrderedDict()

        by_platform: "OrderedDict[str, List[Dict[str, Any]]]" = OrderedDict()
        for row in results:
            platform_name = row[1]
            units = _i(row[4])
            gross_sales = _f(row[5])
            net_sales = _f(row[6])
            units_returned = _i(row[7])
            paid_units = _i(row[8])
            ad_spend = _f(row[9])
            profit = _f(row[10])
            organic_units = max(units - paid_units, 0)

            by_platform.setdefault(platform_name, []).append({
                "sku": row[2],
                "product": row[3],
                "units": units,
                "gross_sales": gross_sales,
                "returns_pct": _pct(units_returned, units),
                "organic_units": organic_units,
                "paid_units": paid_units,
                "ad_spend": ad_spend,
                "tacos_pct": _pct(ad_spend, gross_sales),
                "profit": profit,
                "margin_pct": _pct(profit, net_sales),
            })
        return by_platform

    # -- Sales-per-SKU rollup (feeds inventory position + consolidated SKU) --

    @staticmethod
    def _get_sku_sales_summary(
        db: Session,
        start_date: date,
        end_date: date,
        platform_filter: Optional[str] = None,
    ) -> "OrderedDict[str, Dict[str, Any]]":
        """Cross-platform sales rollup per SKU."""
        query = """
        SELECT
            sku,
            product_name,
            COALESCE(SUM(units_sold), 0) AS units,
            COALESCE(SUM(gross_sales), 0) AS gross_sales,
            COALESCE(SUM(net_sales), 0) AS net_sales,
            COALESCE(SUM(units_returned), 0) AS units_returned,
            COALESCE(SUM(attributed_units), 0) AS paid_units,
            COALESCE(SUM(ad_spend), 0) AS ad_spend,
            COALESCE(SUM(contribution_inr), 0) AS profit
        FROM vw_product_platform_daily
        WHERE date BETWEEN :start_date AND :end_date
        """
        params: Dict[str, Any] = {"start_date": start_date, "end_date": end_date}
        if platform_filter:
            query += " AND platform_id = :platform_filter"
            params["platform_filter"] = platform_filter
        query += " GROUP BY sku, product_name ORDER BY gross_sales DESC"

        try:
            results = db.execute(text(query), params).fetchall()
        except Exception as e:
            logger.error(f"SKU sales summary query failed: {str(e)}", exc_info=True)
            return OrderedDict()

        summary: "OrderedDict[str, Dict[str, Any]]" = OrderedDict()
        for row in results:
            units = _i(row[2])
            gross_sales = _f(row[3])
            net_sales = _f(row[4])
            units_returned = _i(row[5])
            paid_units = _i(row[6])
            ad_spend = _f(row[7])
            profit = _f(row[8])
            organic_units = max(units - paid_units, 0)

            summary[row[0]] = {
                "sku": row[0],
                "product": row[1],
                "units": units,
                "gross_sales": gross_sales,
                "net_sales": net_sales,
                "returns_pct": _pct(units_returned, units),
                "organic_units": organic_units,
                "paid_units": paid_units,
                "ad_spend": ad_spend,
                "tacos_pct": _pct(ad_spend, gross_sales),
                "profit": profit,
                "margin_pct": _pct(profit, net_sales),
            }
        return summary

    # -- Section 3: Inventory & Warehouse -------------------------------

    @staticmethod
    def _get_warehouse_ops_rows(
        db: Session,
        start_date: date,
        end_date: date,
        warehouse_filter: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        """Per warehouse+SKU operational rollup for the period.

        Opening Stock = opening_stock on the first day-in-range;
        Closing Stock / Reorder Level / DOS / Stock Status = the values as
        of the last day-in-range; Inbound/Outbound Qty = summed over the
        whole period (Outbound Qty = fulfilled_units, the closest real
        column to "units shipped out"). "Damaged / Hold" is dropped - there
        is no such column in inventory_daily.
        """
        query = """
        WITH ranked AS (
            SELECT
                idy.warehouse_id,
                idy.sku,
                idy.inventory_date,
                idy.opening_stock,
                idy.inbound_stock,
                idy.fulfilled_units,
                idy.closing_stock,
                idy.reorder_point,
                idy.days_of_cover,
                idy.stock_status,
                ROW_NUMBER() OVER (PARTITION BY idy.warehouse_id, idy.sku ORDER BY idy.inventory_date ASC) AS rn_first,
                ROW_NUMBER() OVER (PARTITION BY idy.warehouse_id, idy.sku ORDER BY idy.inventory_date DESC) AS rn_last
            FROM inventory_daily idy
            WHERE idy.inventory_date BETWEEN :start_date AND :end_date
        """
        params: Dict[str, Any] = {"start_date": start_date, "end_date": end_date}
        if warehouse_filter:
            query += " AND idy.warehouse_id = :warehouse_filter"
            params["warehouse_filter"] = warehouse_filter
        query += """
        )
        SELECT
            r.warehouse_id,
            w.warehouse_name,
            r.sku,
            pr.product_name,
            MAX(CASE WHEN r.rn_first = 1 THEN r.opening_stock END) AS opening_stock,
            SUM(r.inbound_stock) AS inbound_qty,
            SUM(r.fulfilled_units) AS outbound_qty,
            MAX(CASE WHEN r.rn_last = 1 THEN r.closing_stock END) AS closing_stock,
            MAX(CASE WHEN r.rn_last = 1 THEN r.reorder_point END) AS reorder_level,
            MAX(CASE WHEN r.rn_last = 1 THEN r.days_of_cover END) AS dos,
            MAX(CASE WHEN r.rn_last = 1 THEN r.stock_status END) AS stock_status
        FROM ranked r
        INNER JOIN warehouses w ON r.warehouse_id = w.warehouse_id
        INNER JOIN products pr ON r.sku = pr.sku
        GROUP BY r.warehouse_id, w.warehouse_name, r.sku, pr.product_name
        ORDER BY w.warehouse_name, r.sku
        """

        try:
            results = db.execute(text(query), params).fetchall()
        except Exception as e:
            logger.error(f"Warehouse ops query failed: {str(e)}", exc_info=True)
            return []

        rows = []
        for row in results:
            dos = row[9]
            rows.append({
                "warehouse_id": row[0],
                "warehouse": row[1],
                "sku": row[2],
                "product": row[3],
                "opening_stock": _i(row[4]),
                "inbound_qty": _i(row[5]),
                "outbound_qty": _i(row[6]),
                "closing_stock": _i(row[7]),
                "reorder_level": _i(row[8]),
                "dos": _f(dos) if dos is not None else None,
                "stock_status": row[10] or "Unknown",
            })
        return rows

    @staticmethod
    def _get_inventory_position_data(
        sku_sales: "OrderedDict[str, Dict[str, Any]]",
        warehouse_rows: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Per-SKU cross-warehouse total inventory position.

        Combines the sales rollup (All Units, Gross Sales) with the
        warehouse operations rollup (Stock, DOS averaged across warehouses
        actually carrying that SKU).
        """
        by_sku_stock: Dict[str, Dict[str, Any]] = {}
        for row in warehouse_rows:
            sku = row["sku"]
            entry = by_sku_stock.setdefault(sku, {
                "product": row["product"], "stock": 0, "dos_values": [],
            })
            entry["stock"] += row["closing_stock"]
            if row["dos"] is not None:
                entry["dos_values"].append(row["dos"])

        all_skus = OrderedDict()
        for sku, sales in sku_sales.items():
            all_skus[sku] = sales.get("product")
        for sku, stock_entry in by_sku_stock.items():
            all_skus.setdefault(sku, stock_entry.get("product"))

        rows = []
        for sku, product_name in all_skus.items():
            sales = sku_sales.get(sku, {})
            stock_entry = by_sku_stock.get(sku, {})
            dos_values = stock_entry.get("dos_values", [])
            rows.append({
                "sku": sku,
                "product": sales.get("product") or product_name,
                "all_units": sales.get("units", 0),
                "gross_sales": sales.get("gross_sales", 0.0),
                "stock": stock_entry.get("stock", 0),
                "dos": (sum(dos_values) / len(dos_values)) if dos_values else None,
            })
        rows.sort(key=lambda r: r["gross_sales"], reverse=True)
        return rows

    # -- Section 4: Consolidated SKU Performance ------------------------

    @staticmethod
    def _get_consolidated_sku_data(
        sku_sales: "OrderedDict[str, Dict[str, Any]]",
        warehouse_rows: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Cross-channel rollup per SKU, with DOS merged in from inventory."""
        dos_by_sku: Dict[str, List[float]] = defaultdict(list)
        for row in warehouse_rows:
            if row["dos"] is not None:
                dos_by_sku[row["sku"]].append(row["dos"])

        rows = []
        for sku, sales in sku_sales.items():
            dos_values = dos_by_sku.get(sku, [])
            rows.append({
                "sku": sku,
                "product": sales["product"],
                "all_units": sales["units"],
                "gross_sales": sales["gross_sales"],
                "returns_pct": sales["returns_pct"],
                "organic_units": sales["organic_units"],
                "paid_units": sales["paid_units"],
                "ad_spend": sales["ad_spend"],
                "tacos_pct": sales["tacos_pct"],
                "profit": sales["profit"],
                "margin_pct": sales["margin_pct"],
                "dos": (sum(dos_values) / len(dos_values)) if dos_values else None,
            })
        return rows

    # -- Section 5: Consolidated P&L -------------------------------------

    @staticmethod
    def _get_pnl_data(
        db: Session,
        start_date: date,
        end_date: date,
        platform_filter: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Consolidated P&L built from the same cost/sales columns that
        drive vw_product_platform_daily.contribution_inr, so the final
        EBITDA line reconciles exactly with the KPI summary's profit figure.
        """
        query = """
        SELECT
            COALESCE(SUM(gross_sales), 0) AS gross_sales,
            COALESCE(SUM(refund_amount), 0) AS refunds,
            COALESCE(SUM(net_sales), 0) AS net_sales,
            COALESCE(SUM(product_cost), 0) AS cogs,
            COALESCE(SUM(ad_spend), 0) AS ad_spend,
            COALESCE(SUM(platform_fee), 0) AS platform_fee,
            COALESCE(SUM(shipping_cost), 0) AS shipping_cost,
            COALESCE(SUM(payment_fee), 0) AS payment_fee,
            COALESCE(SUM(other_variable_cost), 0) AS other_cost,
            COALESCE(SUM(contribution_inr), 0) AS ebitda
        FROM vw_product_platform_daily
        WHERE date BETWEEN :start_date AND :end_date
        """
        params: Dict[str, Any] = {"start_date": start_date, "end_date": end_date}
        if platform_filter:
            query += " AND platform_id = :platform_filter"
            params["platform_filter"] = platform_filter

        try:
            row = db.execute(text(query), params).fetchone()
        except Exception as e:
            logger.error(f"P&L query failed: {str(e)}", exc_info=True)
            row = None

        if not row:
            return {"line_items": [], "gmv": 0.0, "ebitda": 0.0, "ebitda_margin_pct": 0.0}

        gross_sales = _f(row[0])
        refunds = _f(row[1])
        net_sales = _f(row[2])
        cogs = _f(row[3])
        ad_spend = _f(row[4])
        commission_logistics = _f(row[5]) + _f(row[6]) + _f(row[7]) + _f(row[8])
        ebitda = _f(row[9])
        net_realized = net_sales - refunds

        def item(label, amount, note):
            return {
                "label": label,
                "amount": amount,
                "pct_of_revenue": _pct(amount, gross_sales),
                "note": note,
            }

        line_items = [
            item(
                "Total Gross Sales Turnover (GMV)", gross_sales,
                "Total invoiced sales value across all active channels before deductions.",
            ),
            item(
                "Less: Returns & Customer Refunds", -refunds,
                "Value of units returned and refunded to customers.",
            ),
            item(
                "Net Realized Sales Turnover", net_realized,
                "Net sales after discounts (net sales) and customer refunds.",
            ),
            item(
                "Less: COGS", -cogs,
                "Direct product cost of goods sold.",
            ),
            item(
                "Less: Total Marketing & Ad Spend", -ad_spend,
                "Total advertising spend across all platforms.",
            ),
            item(
                "Less: Marketplace Commission & Logistics", -commission_logistics,
                "Platform fees, shipping, and payment-processing costs.",
            ),
            item(
                "GRAND NET OPERATING PROFIT (EBITDA)", ebitda,
                "Operating profit before interest, tax, depreciation and amortization.",
            ),
        ]

        return {
            "line_items": line_items,
            "gmv": gross_sales,
            "ebitda": ebitda,
            "ebitda_margin_pct": _pct(ebitda, gross_sales),
        }

    # -- Section 6: Audit Notes / Actions (agents/ pipeline) --------------

    @staticmethod
    def _get_agent_metrics_rows(
        db: Session,
        start_date: date,
        end_date: date,
        platform_filter: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        """Full per platform+SKU aggregate needed to build ProductMetrics/
        PlatformMetrics dataclasses for the DataAnalysisAgent /
        InsightRecommendationAgent pipeline. This is a superset of the
        trimmed display columns used in the PDF tables above."""
        query = """
        SELECT
            platform_id,
            platform,
            sku,
            product_name,
            COALESCE(SUM(orders), 0) AS orders,
            COALESCE(SUM(units_sold), 0) AS units_sold,
            COALESCE(SUM(gross_sales), 0) AS gross_sales,
            COALESCE(SUM(discount), 0) AS discount,
            COALESCE(SUM(net_sales), 0) AS net_sales,
            COALESCE(SUM(ad_spend), 0) AS ad_spend,
            COALESCE(SUM(attributed_units), 0) AS ad_attributed_units,
            COALESCE(SUM(ad_attributed_sales), 0) AS ad_attributed_sales,
            COALESCE(SUM(product_cost), 0) AS product_cost,
            COALESCE(SUM(platform_fee), 0) AS platform_fee,
            COALESCE(SUM(shipping_cost), 0) AS shipping_cost,
            COALESCE(SUM(payment_fee), 0) AS payment_fee,
            COALESCE(SUM(other_variable_cost), 0) AS other_cost,
            COALESCE(SUM(units_returned), 0) AS units_returned,
            COALESCE(SUM(refund_amount), 0) AS refund_amount,
            COALESCE(SUM(units_cancelled), 0) AS units_cancelled,
            COALESCE(SUM(contribution_inr), 0) AS contribution
        FROM vw_product_platform_daily
        WHERE date BETWEEN :start_date AND :end_date
        """
        params: Dict[str, Any] = {"start_date": start_date, "end_date": end_date}
        if platform_filter:
            query += " AND platform_id = :platform_filter"
            params["platform_filter"] = platform_filter
        query += " GROUP BY platform_id, platform, sku, product_name"

        try:
            results = db.execute(text(query), params).fetchall()
        except Exception as e:
            logger.error(f"Agent metrics query failed: {str(e)}", exc_info=True)
            return []

        rows = []
        for row in results:
            rows.append({
                "platform_id": row[0], "platform": row[1], "sku": row[2], "product_name": row[3],
                "orders": _i(row[4]), "units_sold": _i(row[5]), "gross_sales": _f(row[6]),
                "discount": _f(row[7]), "net_sales": _f(row[8]), "ad_spend": _f(row[9]),
                "ad_attributed_units": _i(row[10]), "ad_attributed_sales": _f(row[11]),
                "product_cost": _f(row[12]), "platform_fee": _f(row[13]), "shipping_cost": _f(row[14]),
                "payment_fee": _f(row[15]), "other_cost": _f(row[16]), "units_returned": _i(row[17]),
                "refund_amount": _f(row[18]), "units_cancelled": _i(row[19]), "contribution": _f(row[20]),
            })
        return rows

    @staticmethod
    def _build_product_metrics(row: Dict[str, Any]) -> ProductMetrics:
        units_sold = row["units_sold"]
        ad_attr_units = row["ad_attributed_units"]
        organic_units = max(units_sold - ad_attr_units, 0)
        ad_attr_sales = row["ad_attributed_sales"]
        net_sales = row["net_sales"]
        organic_sales = max(net_sales - ad_attr_sales, 0.0)
        ad_spend = row["ad_spend"]
        product_cost = row["product_cost"]
        platform_fee = row["platform_fee"]
        shipping_cost = row["shipping_cost"]
        payment_fee = row["payment_fee"]
        other_cost = row["other_cost"]
        total_cost = product_cost + platform_fee + shipping_cost + payment_fee + other_cost + ad_spend
        units_returned = row["units_returned"]
        units_cancelled = row["units_cancelled"]
        orders = row["orders"]
        contribution = row["contribution"]
        profit_margin_pct = _pct(contribution, net_sales)

        if contribution < 0:
            status = "Critical"
        elif profit_margin_pct < 15:
            status = "At Risk"
        else:
            status = "Healthy"

        return ProductMetrics(
            sku=row["sku"],
            product_name=row["product_name"],
            units_sold=units_sold,
            gross_sales_inr=row["gross_sales"],
            net_sales_inr=net_sales,
            discount_inr=row["discount"],
            ad_spend_inr=ad_spend,
            ad_attributed_units=ad_attr_units,
            ad_attributed_sales_inr=ad_attr_sales,
            organic_units=organic_units,
            organic_sales_inr=organic_sales,
            organic_share_pct=_pct(organic_units, units_sold),
            roas=_ratio(ad_attr_sales, ad_spend),
            acos_pct=_pct(ad_spend, ad_attr_sales),
            product_cost_inr=product_cost,
            platform_fee_inr=platform_fee,
            shipping_cost_inr=shipping_cost,
            payment_fee_inr=payment_fee,
            other_cost_inr=other_cost,
            total_cost_inr=total_cost,
            units_returned=units_returned,
            refund_amount_inr=row["refund_amount"],
            return_rate_pct=_pct(units_returned, units_sold),
            units_cancelled=units_cancelled,
            cancellation_rate_pct=_pct(units_cancelled, orders),
            contribution_inr=contribution,
            profit_margin_pct=profit_margin_pct,
            profitability_status=status,
        )

    @staticmethod
    def _build_platform_metrics(platform_id: str, platform_rows: List[Dict[str, Any]]) -> PlatformMetrics:
        totals = defaultdict(float)
        sku_seen = set()
        top_sku, top_sales = None, -1.0
        platform_name = platform_rows[0]["platform"] if platform_rows else platform_id

        for row in platform_rows:
            for key in (
                "orders", "units_sold", "gross_sales", "discount", "net_sales", "ad_spend",
                "ad_attributed_units", "ad_attributed_sales", "product_cost", "platform_fee",
                "shipping_cost", "payment_fee", "other_cost", "units_returned", "refund_amount",
                "units_cancelled", "contribution",
            ):
                totals[key] += row[key]
            organic_units = max(row["units_sold"] - row["ad_attributed_units"], 0)
            totals["organic_units"] += organic_units
            totals["organic_sales"] += max(row["net_sales"] - row["ad_attributed_sales"], 0.0)
            totals["total_cost"] += (
                row["product_cost"] + row["platform_fee"] + row["shipping_cost"]
                + row["payment_fee"] + row["other_cost"] + row["ad_spend"]
            )
            sku_seen.add(row["sku"])
            if row["gross_sales"] > top_sales:
                top_sales, top_sku = row["gross_sales"], row["sku"]

        net_sales = totals["net_sales"]
        contribution = totals["contribution"]

        return PlatformMetrics(
            platform_id=platform_id,
            platform_name=platform_name,
            total_orders=int(totals["orders"]),
            total_units_sold=int(totals["units_sold"]),
            total_gross_sales_inr=totals["gross_sales"],
            total_net_sales_inr=net_sales,
            total_discount_inr=totals["discount"],
            total_ad_spend_inr=totals["ad_spend"],
            total_ad_attributed_units=int(totals["ad_attributed_units"]),
            total_ad_attributed_sales_inr=totals["ad_attributed_sales"],
            total_organic_units=int(totals["organic_units"]),
            total_organic_sales_inr=totals["organic_sales"],
            platform_roas=_ratio(totals["ad_attributed_sales"], totals["ad_spend"]),
            platform_acos_pct=_pct(totals["ad_spend"], totals["ad_attributed_sales"]),
            total_product_cost_inr=totals["product_cost"],
            total_platform_fee_inr=totals["platform_fee"],
            total_shipping_cost_inr=totals["shipping_cost"],
            total_payment_fee_inr=totals["payment_fee"],
            total_other_cost_inr=totals["other_cost"],
            total_cost_inr=totals["total_cost"],
            total_returns=int(totals["units_returned"]),
            total_refund_inr=totals["refund_amount"],
            overall_return_rate_pct=_pct(totals["units_returned"], totals["units_sold"]),
            total_cancellations=int(totals["units_cancelled"]),
            overall_cancellation_rate_pct=_pct(totals["units_cancelled"], totals["orders"]),
            total_contribution_inr=contribution,
            overall_profit_margin_pct=_pct(contribution, net_sales),
            product_count=len(sku_seen),
            top_product_sku=top_sku,
            top_product_sales_inr=top_sales if top_sku else None,
        )

    @staticmethod
    def _derive_due_date(generated_at: date, timeline: Optional[str]) -> str:
        """Best-effort due date derived from a recommendation's `timeline`
        text (e.g. "Within 30 days", "Immediate"). Left blank if the
        timeline text doesn't parse into a concrete offset."""
        if not timeline:
            return ""
        t = timeline.lower()
        if "immediate" in t:
            return generated_at.isoformat()
        m = re.search(r"(\d+)\s*day", t)
        if m:
            return (generated_at + timedelta(days=int(m.group(1)))).isoformat()
        m = re.search(r"(\d+)\s*week", t)
        if m:
            return (generated_at + timedelta(weeks=int(m.group(1)))).isoformat()
        return ""

    @staticmethod
    def _get_audit_notes(
        db: Session,
        start_date: date,
        end_date: date,
        platform_filter: Optional[str] = None,
    ) -> Tuple[List[Dict[str, Any]], str]:
        """Run real DB metrics through the actual agents/ pipeline
        (DataAnalysisAgent -> InsightRecommendationAgent) to produce the
        Audit Notes / Actions section. Returns (rows, source) where source
        records whether the full agent wrapper ran or a fallback was used.
        """
        try:
            agent_rows = ReportService._get_agent_metrics_rows(db, start_date, end_date, platform_filter)
            if not agent_rows:
                return [], "agents (no data in period)"

            product_metrics = [ReportService._build_product_metrics(r) for r in agent_rows]

            by_platform: "OrderedDict[str, List[Dict[str, Any]]]" = OrderedDict()
            for row in agent_rows:
                by_platform.setdefault(row["platform_id"], []).append(row)
            platform_metrics = [
                ReportService._build_platform_metrics(pid, rows) for pid, rows in by_platform.items()
            ]

            analysis_agent = DataAnalysisAgent()
            findings = []
            for pm in platform_metrics:
                findings.extend(analysis_agent.analyze_platform_performance(pm))
            for prod_m in product_metrics:
                findings.extend(analysis_agent.analyze_product_performance(prod_m))

            anomalies = analysis_agent.detect_anomalies(list(platform_metrics) + list(product_metrics))

            analysis_result = analysis_agent.generate_analysis_result(
                period_start=start_date,
                period_end=end_date,
                analysis_type="channel_and_product_performance",
                findings=findings,
                anomalies=anomalies,
                key_metrics={
                    "platforms_analyzed": len(platform_metrics),
                    "products_analyzed": len(product_metrics),
                },
            )

            ir_agent = InsightRecommendationAgent()
            result = ir_agent.analyze(
                analysis_result,
                data_completeness=1.0,
                generated_at=date.today(),
            )

            priority_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
            recommendations = sorted(
                result.recommendations,
                key=lambda r: priority_order.get(getattr(r.priority, "value", str(r.priority)), 5),
            )

            insights_by_id = {i.insight_id: i for i in result.insights}

            rows = []
            seen = set()
            for rec in recommendations:
                observation = None
                for insight_id in rec.insight_sources:
                    if insight_id in insights_by_id:
                        observation = insights_by_id[insight_id].description
                        break
                if not observation:
                    observation = rec.rationale

                scope = rec.product_name or rec.platform_name or ""
                action_owner = f"{rec.action} (Owner: {rec.owner})"
                if scope:
                    action_owner = f"[{scope}] {action_owner}"

                # The same SKU can surface an identical-looking statistical
                # anomaly on more than one platform (the agent's anomaly
                # description doesn't carry a platform qualifier), which
                # would otherwise duplicate rows here; dedupe on the
                # rendered (observation, action) pair.
                dedupe_key = (observation, action_owner)
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)

                rows.append({
                    "observation": observation,
                    "action_owner": action_owner,
                    "due_date": ReportService._derive_due_date(date.today(), rec.timeline),
                    "status": "Open",
                    "priority": getattr(rec.priority, "value", str(rec.priority)),
                })

                if len(rows) >= 25:
                    break

            if not rows:
                rows.append({
                    "observation": result.management_summary.executive_summary,
                    "action_owner": "Continue monitoring performance across channels",
                    "due_date": "",
                    "status": "Informational",
                    "priority": "info",
                })

            return rows, "agents (DataAnalysisAgent + InsightRecommendationAgent)"

        except Exception as e:
            logger.error(f"Agent audit-notes pipeline failed: {str(e)}", exc_info=True)
            return [{
                "observation": "Automated audit-notes pipeline could not run for this period.",
                "action_owner": f"Investigate report generation error: {str(e)[:200]}",
                "due_date": "",
                "status": "Open",
                "priority": "high",
            }], "fallback (agent pipeline raised an exception)"

    # -- Legacy sections (advertising / profitability report types) ------

    @staticmethod
    def _get_profitability_trend(db: Session, start_date: date, end_date: date) -> Dict[str, Any]:
        """Daily profitability trend, alongside the new consolidated P&L."""
        query = """
        SELECT
            DATE(date) as day,
            SUM(total_gross_sales) as revenue,
            SUM(total_contribution) as profit,
            AVG(overall_profit_margin_pct) as margin
        FROM vw_daily_kpi_summary
        WHERE date BETWEEN :start_date AND :end_date
        GROUP BY DATE(date)
        ORDER BY day ASC
        """

        try:
            results = db.execute(text(query), {"start_date": start_date, "end_date": end_date}).fetchall()
        except Exception as e:
            logger.error(f"Profitability trend query failed: {str(e)}", exc_info=True)
            results = []

        daily_data = [
            {
                "date": str(row[0]),
                "revenue": _f(row[1]),
                "profit": _f(row[2]),
                "margin": _f(row[3]),
            }
            for row in results
        ]

        return {
            "daily_trends": daily_data,
            "summary": {
                "avg_margin": sum(d["margin"] for d in daily_data) / len(daily_data) if daily_data else 0,
                "total_profit": sum(d["profit"] for d in daily_data),
                "total_revenue": sum(d["revenue"] for d in daily_data),
            }
        }

    @staticmethod
    def _get_advertising_data(db: Session, start_date: date, end_date: date) -> Dict[str, Any]:
        """Get advertising analytics data."""
        query = """
        SELECT
            DATE(date) as day,
            SUM(total_ad_spend) as spend,
            SUM(total_ad_sales) as attributed_sales,
            AVG(overall_roas) as roas
        FROM vw_daily_kpi_summary
        WHERE date BETWEEN :start_date AND :end_date
        GROUP BY DATE(date)
        ORDER BY day ASC
        """

        try:
            results = db.execute(text(query), {"start_date": start_date, "end_date": end_date}).fetchall()
        except Exception as e:
            logger.error(f"Advertising data query failed: {str(e)}", exc_info=True)
            results = []

        daily_data = [
            {
                "date": str(row[0]),
                "spend": _f(row[1]),
                "attributed_sales": _f(row[2]),
                "roas": _f(row[3]),
            }
            for row in results
        ]

        return {
            "daily_trends": daily_data,
            "summary": {
                "total_spend": sum(d["spend"] for d in daily_data),
                "total_attributed_sales": sum(d["attributed_sales"] for d in daily_data),
                "avg_roas": sum(d["roas"] for d in daily_data) / len(daily_data) if daily_data else 0,
            }
        }

    # ------------------------------------------------------------------
    # Report content assembly
    # ------------------------------------------------------------------

    @staticmethod
    def _generate_report_content(
        report_id: str,
        report_type: str,
        report_title: str,
        data: Dict[str, Any],
        start_date: date,
        end_date: date,
        include_recommendations: bool = True,
    ) -> Dict[str, Any]:
        """Generate report content with insights and recommendations."""

        content = {
            "report_id": report_id,
            "title": report_title,
            "type": report_type,
            "period": f"{start_date} to {end_date}",
            "generated_at": datetime.now().isoformat(),
            "data": data,
        }

        if include_recommendations and "audit_notes" not in data:
            content["recommendations"] = ReportService._get_recommendations(report_type, data)

        return content

    @staticmethod
    def _get_recommendations(report_type: str, data: Dict[str, Any]) -> List[str]:
        """Generate actionable recommendations based on report data.

        Only used for report types that don't already carry agent-derived
        audit notes (executive_summary / management_monthly use the
        agents/ pipeline instead - see `_get_audit_notes`).
        """

        recommendations = []

        if report_type == "platform_analysis" and data.get("channel_data"):
            top = data["channel_data"][0]
            recommendations.append(f"Focus on {top['platform']} - it has the highest gross sales")

        elif report_type == "product_analysis" and data.get("consolidated_sku"):
            top = max(data["consolidated_sku"], key=lambda p: p["profit"], default=None)
            if top:
                recommendations.append(f"Scale {top['product']} - it's your profit leader")

        elif report_type == "advertising" and "advertising_data" in data:
            ad_data = data["advertising_data"]
            if ad_data.get("summary", {}).get("avg_roas", 0) < 3:
                recommendations.append("Review ad creative and targeting - ROAS is below 3x")

        return recommendations

    # ------------------------------------------------------------------
    # PDF rendering
    # ------------------------------------------------------------------

    @staticmethod
    def _render_pdf(report_content: Dict[str, Any], file_path: str) -> None:
        """Render the canonical report dictionary as a readable, multi-
        section PDF mirroring the Sleepsia Omni-Channel Audit template
        (adapted to whatever sections are actually present in the data).
        """
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
        )
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT
        from xml.sax.saxutils import escape as _xml_escape

        HEADER_BG = colors.HexColor("#1f2d3d")
        ALT_ROW_BG = colors.HexColor("#f2f4f7")
        BORDER = colors.HexColor("#c9ced6")
        SECTION_COLOR = colors.HexColor("#1f2d3d")

        data = report_content.get("data", {})
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "ReportTitle", parent=styles["Title"], fontSize=20, textColor=HEADER_BG,
        )
        section_style = ParagraphStyle(
            "SectionHeading", parent=styles["Heading2"], fontSize=13,
            textColor=SECTION_COLOR, spaceBefore=14, spaceAfter=8, fontName="Helvetica-Bold",
        )
        sub_style = ParagraphStyle(
            "SubHeading", parent=styles["Heading3"], fontSize=10.5,
            textColor=colors.HexColor("#40495a"), spaceBefore=8, spaceAfter=4,
        )
        note_style = ParagraphStyle(
            "Note", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#666666"),
        )
        cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7.6, leading=9, alignment=TA_LEFT)
        header_cell_style = ParagraphStyle(
            "HeaderCell", parent=styles["Normal"], fontSize=7.8, leading=9,
            textColor=colors.white, fontName="Helvetica-Bold",
        )

        def esc(text_value: Any) -> str:
            """Escape for reportlab's Paragraph mini-XML parser (it treats
            '&', '<', '>' as markup, so raw text - e.g. "P&L", product
            names - must be escaped or it renders mangled/truncated)."""
            return _xml_escape("" if text_value is None else str(text_value))

        def money(v) -> str:
            # reportlab's base Helvetica font has no glyph for U+20B9 (₹),
            # which renders as a black box; "Rs. " is guaranteed to render.
            v = _f(v)
            sign = "-" if v < 0 else ""
            return f"{sign}Rs. {abs(v):,.0f}"

        def num(v) -> str:
            return f"{_f(v):,.0f}"

        def pct(v) -> str:
            return f"{_f(v):.1f}%"

        def dos_fmt(v) -> str:
            return f"{_f(v):.1f}" if v is not None else "N/A"

        def wrap(text_value) -> Paragraph:
            return Paragraph(esc(text_value), cell_style)

        def P(text_value, style) -> Paragraph:
            """Escaped Paragraph helper for section/sub headings and any
            other free text that isn't already going through wrap()."""
            return Paragraph(esc(text_value), style)

        def styled_table(rows: List[List[Any]], col_widths: List[float]) -> Table:
            header = [Paragraph(esc(c), header_cell_style) for c in rows[0]]
            body = [header] + rows[1:]
            table = Table(body, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), HEADER_BG),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, BORDER),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ALT_ROW_BG]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))
            return table

        story: List[Any] = []

        # ---- Cover -----------------------------------------------------
        story.append(P(report_content.get("title", "Sleepsia Report"), title_style))
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            f"Period: {report_content.get('period', 'Not specified')} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Report ID: {report_content.get('report_id', '')} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}",
            styles["Normal"],
        ))
        story.append(Spacer(1, 10))

        kpi = data.get("kpi_summary", {})
        if kpi:
            kpi_rows = [["Metric", "Value"], [
                "Gross Revenue", money(kpi.get("revenue", 0))],
                ["Net Revenue", money(kpi.get("net_revenue", 0))],
                ["Profit (Contribution)", money(kpi.get("profit", 0))],
                ["Profit Margin", pct(kpi.get("profit_margin_pct", 0))],
                ["Orders", num(kpi.get("orders", 0))],
                ["Units Sold", num(kpi.get("units_sold", 0))],
                ["Ad Spend", money(kpi.get("ad_spend", 0))],
                ["ROAS", f"{_f(kpi.get('roas', 0)):.2f}x"],
            ]
            story.append(Paragraph("KPI Summary", section_style))
            story.append(styled_table(kpi_rows, [2.6 * inch, 2.2 * inch]))
            story.append(Spacer(1, 10))

        # ---- Section 1: Channel Performance -----------------------------
        channel_data = data.get("channel_data")
        if channel_data:
            story.append(Paragraph("1. Channel Performance", section_style))
            rows = [["Platform", "Orders", "Units", "Gross Sales", "Returns", "Net Revenue",
                     "Ad Spend", "TACoS%", "Profit", "Margin%"]]
            for c in channel_data:
                rows.append([
                    wrap(c["platform"]), num(c["orders"]), num(c["units"]), money(c["gross_sales"]),
                    money(c["returns_amount"]), money(c["net_revenue"]), money(c["ad_spend"]),
                    pct(c["tacos_pct"]), money(c["profit"]), pct(c["margin_pct"]),
                ])
            widths = [1.3, 0.75, 0.75, 1.05, 0.95, 1.05, 0.95, 0.7, 1.0, 0.75]
            story.append(styled_table(rows, [w * inch for w in widths]))
            story.append(Spacer(1, 10))

        # ---- Section 2: Product Performance by Platform ------------------
        product_by_platform = data.get("product_by_platform")
        if product_by_platform:
            story.append(Paragraph("2. Product Performance by Platform", section_style))
            widths = [0.62, 1.55, 0.6, 0.95, 0.7, 0.62, 0.55, 0.85, 0.62, 0.85, 0.65]
            for platform_name, products in product_by_platform.items():
                story.append(P(platform_name, sub_style))
                rows = [["SKU", "Product", "Units", "Gross Sales", "Returns%", "Organic",
                         "Paid", "Ad Spend", "TACoS%", "Profit", "Margin%"]]
                for p in products:
                    rows.append([
                        wrap(p["sku"]), wrap(p["product"]), num(p["units"]), money(p["gross_sales"]),
                        pct(p["returns_pct"]), num(p["organic_units"]), num(p["paid_units"]),
                        money(p["ad_spend"]), pct(p["tacos_pct"]), money(p["profit"]), pct(p["margin_pct"]),
                    ])
                story.append(styled_table(rows, [w * inch for w in widths]))
                story.append(Spacer(1, 8))

        # ---- Section 3: Inventory & Warehouse -----------------------------
        inventory_position = data.get("inventory_position")
        warehouse_operations = data.get("warehouse_operations")
        if inventory_position or warehouse_operations:
            story.append(PageBreak())
            story.append(P("3. Inventory & Warehouse", section_style))

            if inventory_position:
                story.append(Paragraph("Inventory Position (per SKU, cross-warehouse total)", sub_style))
                rows = [["SKU", "Product", "All Units", "Gross Sales", "Stock / Inventory", "DOS"]]
                for r in inventory_position:
                    rows.append([
                        wrap(r["sku"]), wrap(r["product"]), num(r["all_units"]), money(r["gross_sales"]),
                        num(r["stock"]), dos_fmt(r["dos"]),
                    ])
                widths = [0.75, 2.4, 1.0, 1.25, 1.15, 0.7]
                story.append(styled_table(rows, [w * inch for w in widths]))
                story.append(Spacer(1, 10))

            if warehouse_operations:
                story.append(Paragraph("Warehouse / Inventory Operations (per warehouse + SKU)", sub_style))
                rows = [["Warehouse", "SKU", "Product", "Opening Stock", "Inbound Qty", "Outbound Qty",
                         "Closing Stock", "Reorder Level", "DOS", "Stock Status"]]
                for r in warehouse_operations:
                    rows.append([
                        wrap(r["warehouse"]), wrap(r["sku"]), wrap(r["product"]), num(r["opening_stock"]),
                        num(r["inbound_qty"]), num(r["outbound_qty"]), num(r["closing_stock"]),
                        num(r["reorder_level"]), dos_fmt(r["dos"]), wrap(r["stock_status"]),
                    ])
                widths = [1.25, 0.6, 1.5, 0.85, 0.8, 0.85, 0.85, 0.85, 0.55, 0.95]
                story.append(styled_table(rows, [w * inch for w in widths]))
                story.append(Spacer(1, 10))

        # ---- Section 4: Consolidated SKU Performance -----------------------
        consolidated_sku = data.get("consolidated_sku")
        if consolidated_sku:
            story.append(PageBreak())
            story.append(Paragraph("4. Consolidated SKU Performance", section_style))
            rows = [["SKU", "Product", "All Units", "Gross Sales", "Returns%", "Organic", "Paid",
                     "Ad Spend", "TACoS%", "Profit", "Margin%", "DOS"]]
            for r in consolidated_sku:
                rows.append([
                    wrap(r["sku"]), wrap(r["product"]), num(r["all_units"]), money(r["gross_sales"]),
                    pct(r["returns_pct"]), num(r["organic_units"]), num(r["paid_units"]),
                    money(r["ad_spend"]), pct(r["tacos_pct"]), money(r["profit"]), pct(r["margin_pct"]),
                    dos_fmt(r["dos"]),
                ])
            widths = [0.55, 1.5, 0.65, 0.95, 0.65, 0.55, 0.5, 0.8, 0.6, 0.8, 0.6, 0.5]
            story.append(styled_table(rows, [w * inch for w in widths]))
            story.append(Spacer(1, 10))

        # ---- Section 5: Consolidated P&L -----------------------------------
        pnl = data.get("pnl")
        if pnl and pnl.get("line_items"):
            story.append(PageBreak())
            story.append(P("5. Consolidated P&L", section_style))
            rows = [["Line Item", "Amount (INR)", "% of Revenue", "Note"]]
            for item in pnl["line_items"]:
                is_total = "GRAND NET" in item["label"]
                escaped_label = esc(item["label"])
                label = f"<b>{escaped_label}</b>" if is_total else escaped_label
                rows.append([
                    Paragraph(label, cell_style), money(item["amount"]),
                    pct(item["pct_of_revenue"]), wrap(item["note"]),
                ])
            widths = [2.5, 1.3, 1.0, 4.2]
            table = styled_table(rows, [w * inch for w in widths])
            story.append(table)
            story.append(Spacer(1, 10))

        legacy_profitability = data.get("profitability_data")
        if legacy_profitability and legacy_profitability.get("daily_trends"):
            story.append(Paragraph("Daily Profitability Trend", sub_style))
            rows = [["Date", "Revenue", "Profit", "Margin%"]]
            for d in legacy_profitability["daily_trends"]:
                rows.append([wrap(d["date"]), money(d["revenue"]), money(d["profit"]), pct(d["margin"])])
            story.append(styled_table(rows, [1.3 * inch, 1.4 * inch, 1.4 * inch, 1.0 * inch]))
            story.append(Spacer(1, 10))

        legacy_advertising = data.get("advertising_data")
        if legacy_advertising and legacy_advertising.get("daily_trends"):
            story.append(Paragraph("Advertising Performance", section_style))
            summary = legacy_advertising.get("summary", {})
            rows = [
                ["Metric", "Value"],
                ["Total Ad Spend", money(summary.get("total_spend", 0))],
                ["Attributed Sales", money(summary.get("total_attributed_sales", 0))],
                ["Average ROAS", f"{_f(summary.get('avg_roas', 0)):.2f}x"],
            ]
            story.append(styled_table(rows, [2.6 * inch, 2.2 * inch]))
            story.append(Spacer(1, 10))

        # ---- Section 6: Audit Notes / Actions --------------------------------
        audit_notes = data.get("audit_notes")
        if audit_notes:
            story.append(PageBreak())
            story.append(Paragraph("6. Audit Notes / Actions", section_style))
            source = data.get("audit_notes_source", "")
            if source:
                story.append(P(f"Generated by: {source}", note_style))
                story.append(Spacer(1, 4))
            rows = [["Observation / Issue", "Action / Owner", "Due Date", "Status"]]
            for n in audit_notes:
                rows.append([
                    wrap(n["observation"]), wrap(n["action_owner"]), wrap(n.get("due_date") or "-"),
                    wrap(n["status"]),
                ])
            widths = [4.0, 3.8, 1.0, 0.85]
            story.append(styled_table(rows, [w * inch for w in widths]))
            story.append(Spacer(1, 10))

        # ---- Free-text recommendations (legacy report types only) -----------
        recommendations = report_content.get("recommendations")
        if recommendations:
            story.append(Paragraph("Recommendations", section_style))
            for rec in recommendations:
                story.append(P(f"• {rec}", styles["Normal"]))
            story.append(Spacer(1, 10))

        story.append(Spacer(1, 6))
        story.append(Paragraph(
            "This report contains confidential Sleepsia business information, generated automatically "
            "from live operational data. Figures with no underlying data source in this period are omitted.",
            note_style,
        ))

        document = SimpleDocTemplate(
            file_path,
            pagesize=landscape(letter),
            leftMargin=0.45 * inch, rightMargin=0.45 * inch,
            topMargin=0.4 * inch, bottomMargin=0.4 * inch,
        )
        document.build(story)

    # ------------------------------------------------------------------
    # Excel rendering
    # ------------------------------------------------------------------

    @staticmethod
    def _render_excel(report_content: Dict[str, Any], file_path: str) -> None:
        """Render the canonical report dictionary as an Excel workbook,
        one sheet per section present in the data."""
        from openpyxl import Workbook

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        summary_sheet.append(["Metric", "Value"])
        data = report_content.get("data", {})
        for key, value in data.get("kpi_summary", {}).items():
            summary_sheet.append([key, value])

        def add_sheet(name: str, rows: List[Dict[str, Any]]):
            if not rows:
                return
            sheet = workbook.create_sheet(name[:31])
            headers = list(rows[0].keys())
            sheet.append(headers)
            for row in rows:
                sheet.append([row.get(h) for h in headers])

        add_sheet("Channel Performance", data.get("channel_data", []))

        product_by_platform = data.get("product_by_platform") or {}
        for platform_name, products in product_by_platform.items():
            add_sheet(f"Products-{platform_name}", products)

        add_sheet("Inventory Position", data.get("inventory_position", []))
        add_sheet("Warehouse Operations", data.get("warehouse_operations", []))
        add_sheet("Consolidated SKU", data.get("consolidated_sku", []))

        pnl = data.get("pnl") or {}
        if pnl.get("line_items"):
            sheet = workbook.create_sheet("Consolidated P&L")
            sheet.append(["Line Item", "Amount (INR)", "% of Revenue", "Note"])
            for item in pnl["line_items"]:
                sheet.append([item["label"], item["amount"], item["pct_of_revenue"], item["note"]])

        add_sheet("Audit Notes", data.get("audit_notes", []))

        legacy_profitability = data.get("profitability_data", {})
        add_sheet("Profitability Trend", legacy_profitability.get("daily_trends", []))

        legacy_advertising = data.get("advertising_data", {})
        add_sheet("Advertising Trend", legacy_advertising.get("daily_trends", []))

        workbook.save(file_path)

    # ------------------------------------------------------------------
    # Report management (list / get / download / email / delete)
    # ------------------------------------------------------------------

    @staticmethod
    def list_reports(db: Session, limit: int = 10, offset: int = 0) -> Dict[str, Any]:
        """List generated reports."""

        ReportService._ensure_reports_dir()

        # Get list of report files
        reports = []
        if os.path.exists(ReportService.REPORTS_DIR):
            for filename in sorted(os.listdir(ReportService.REPORTS_DIR), reverse=True)[offset:offset + limit]:
                if filename.endswith('.json'):
                    file_path = os.path.join(ReportService.REPORTS_DIR, filename)
                    try:
                        with open(file_path, 'r') as f:
                            report_data = json.load(f)
                        structured_period = report_data.get("data", {}).get("period", {})
                        reports.append({
                            "report_id": report_data.get("report_id", filename[:-5]),
                            "report_type": report_data.get("type", "unknown"),
                            "created_at": report_data.get("generated_at", ""),
                            "start_date": structured_period.get("start_date", ""),
                            "end_date": structured_period.get("end_date", ""),
                            "status": "completed",
                            "file_size": os.path.getsize(file_path),
                        })
                    except Exception:
                        pass

        return {
            "reports": reports,
            "total": len(reports),
        }

    @staticmethod
    def get_report(db: Session, report_id: str) -> Optional[Dict[str, Any]]:
        """Get specific report."""

        file_path = os.path.join(ReportService.REPORTS_DIR, f"{report_id}.json")

        if os.path.exists(file_path):
            with open(file_path, 'r') as f:
                report_data = json.load(f)

            # "period" at the top level is a human-readable string
            # (e.g. "2026-07-25 to 2026-08-21"); the structured start/end
            # dates live under data.period, set by _collect_report_data.
            structured_period = report_data.get("data", {}).get("period", {})

            return {
                "report_id": report_data.get("report_id"),
                "report_type": report_data.get("type"),
                "created_at": report_data.get("generated_at"),
                "start_date": structured_period.get("start_date"),
                "end_date": structured_period.get("end_date"),
                "status": "completed",
                "file_size": os.path.getsize(file_path),
                "download_url": f"/api/reports/{report_id}/download",
            }

        return None

    @staticmethod
    def get_report_file(db: Session, report_id: str, format: Optional[str] = None) -> Optional[str]:
        """Get path to report file."""

        file_format = format or "json"
        file_path = os.path.join(ReportService.REPORTS_DIR, f"{report_id}.{file_format}")

        if os.path.exists(file_path):
            return file_path

        # Try default json
        json_path = os.path.join(ReportService.REPORTS_DIR, f"{report_id}.json")
        if os.path.exists(json_path):
            return json_path

        return None

    @staticmethod
    def email_report(
        db: Session,
        report_id: str,
        email_to: str,
        cc: Optional[str] = None,
        bcc: Optional[str] = None,
        format: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Email a report to recipients.

        Defaults to the PDF rendering (the polished, human-readable report)
        rather than the raw JSON - a report email should carry the same
        document the user would have downloaded from the Reports page.
        """
        from automation.email_service import ReportEmailService

        try:
            # Get the report file - prefer PDF unless the caller asked otherwise.
            file_path = ReportService.get_report_file(db, report_id, format or "pdf")
            if not file_path:
                return {
                    "success": False,
                    "error": f"Report {report_id} not found",
                    "report_id": report_id,
                }

            actual_format = file_path.rsplit('.', 1)[-1].lower()

            # Read the report file
            with open(file_path, 'rb') as f:
                report_bytes = f.read()

            # Load report metadata for a more informative subject/body.
            report_meta = ReportService.get_report(db, report_id) or {}
            report_title = ReportService.REPORT_TYPES.get(
                report_meta.get("report_type", ""), "Business Report"
            )
            period_start = report_meta.get("start_date", "")
            period_end = report_meta.get("end_date", "")

            # Parse recipients
            recipients = [email_to]
            cc_list = cc.split(',') if cc else None
            bcc_list = bcc.split(',') if bcc else None

            # Send via email service
            email_service = ReportEmailService()
            success = email_service.send_report(
                subject=f"Sleepsia {report_title} - {period_start} to {period_end}",
                body=f"""
Dear Recipient,

Please find attached your requested Sleepsia business report: {report_title}
Period: {period_start} to {period_end}
Report ID: {report_id}

Best regards,
Sleepsia Analytics System
                """.strip(),
                recipients=recipients,
                cc=cc_list,
                bcc=bcc_list,
                attachments={f"{report_id}.{actual_format}": report_bytes},
            )

            return {
                "success": success,
                "message": f"Report {report_id} sent to {email_to}" if success else f"Failed to send report {report_id}",
                "report_id": report_id,
                "email_to": email_to,
                "timestamp": datetime.now().isoformat(),
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "report_id": report_id,
                "timestamp": datetime.now().isoformat(),
            }

    @staticmethod
    def delete_report(db: Session, report_id: str) -> bool:
        """Delete a report file."""

        file_path = os.path.join(ReportService.REPORTS_DIR, f"{report_id}.json")

        if os.path.exists(file_path):
            try:
                os.remove(file_path)
                return True
            except Exception:
                return False

        return False
