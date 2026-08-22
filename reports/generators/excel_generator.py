"""
Excel Report Generator

Generates comprehensive Excel workbooks from OmniChannelReport data.

Uses openpyxl to create multiple worksheets including:
- Cover sheet with metadata
- Executive summary
- Platform performance (one sheet per platform)
- Consolidated product analysis
- P&L statement
- Channel efficiency
- Recommendations and alerts

The Excel maintains the same data structure as the PDF but in a spreadsheet format
that allows for further analysis and filtering.
"""

import io
from datetime import datetime
from decimal import Decimal
from typing import Optional

from reports.models.report_models import OmniChannelReport
from reports.utils.formatting import (
    format_currency,
    format_percentage,
    format_roas,
    format_units,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
        Border,
        Side,
        numbers,
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelReportGenerator:
    """Generate Excel workbooks from report data."""

    def __init__(self, report_data: OmniChannelReport):
        """
        Initialize Excel generator with report data.

        Args:
            report_data: OmniChannelReport object
        """
        self.report_data = report_data
        self.workbook = None

    def generate(self) -> bytes:
        """
        Generate complete Excel workbook.

        Returns:
            Excel file content as bytes
        """
        if not HAS_OPENPYXL:
            return self._generate_fallback()

        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

        # Create worksheets
        self._create_cover_sheet()
        self._create_executive_summary_sheet()
        self._create_platform_sheets()
        self._create_consolidated_products_sheet()
        self._create_pnl_sheet()
        self._create_channel_efficiency_sheet()
        self._create_recommendations_sheet()

        # Save to bytes
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        excel_content = buffer.getvalue()
        buffer.close()

        return excel_content

    def _create_cover_sheet(self):
        """Create cover/metadata sheet."""
        ws = self.workbook.create_sheet("Cover", 0)
        ws.column_dimensions["A"].width = 50

        title_fill = PatternFill(start_color="1a3a52", end_color="1a3a52", fill_type="solid")
        title_font = Font(bold=True, size=14, color="FFFFFF")

        row = 1
        self._add_cell(ws, row, 1, self.report_data.metadata.report_type, title_font, title_fill)
        row += 2

        self._add_cell(ws, row, 1, "Organization:", Font(bold=True))
        row += 1
        self._add_cell(ws, row, 1, self.report_data.metadata.organization)
        row += 2

        self._add_cell(ws, row, 1, "Audit Date:", Font(bold=True))
        row += 1
        self._add_cell(ws, row, 1, self.report_data.metadata.audit_date.strftime("%d %b %Y"))
        row += 2

        self._add_cell(ws, row, 1, "Report Period:", Font(bold=True))
        row += 1
        if self.report_data.metadata.report_period_start:
            period_text = f"{self.report_data.metadata.report_period_start.strftime('%d %b %Y')} to {self.report_data.metadata.report_period_end.strftime('%d %b %Y')}"
            self._add_cell(ws, row, 1, period_text)
        row += 2

        self._add_cell(ws, row, 1, "Scope:", Font(bold=True))
        row += 1
        self._add_cell(ws, row, 1, self.report_data.metadata.scope)
        row += 2

        self._add_cell(ws, row, 1, "Status:", Font(bold=True))
        row += 1
        self._add_cell(ws, row, 1, self.report_data.metadata.status)
        row += 2

        self._add_cell(ws, row, 1, "Generated:", Font(bold=True))
        row += 1
        self._add_cell(ws, row, 1, self.report_data.metadata.generated_at.strftime("%d %b %Y %H:%M:%S"))

    def _create_executive_summary_sheet(self):
        """Create executive summary sheet."""
        ws = self.workbook.create_sheet("Executive Summary", 1)
        ws.column_dimensions["A"].width = 100

        row = 1

        # Summary text
        self._add_cell(ws, row, 1, "Summary", Font(bold=True, size=12))
        row += 1

        if self.report_data.management_summary:
            self._add_wrapped_cell(ws, row, 1, self.report_data.management_summary.summary_text)
            row += 3

            # Key Findings
            self._add_cell(ws, row, 1, "Key Findings", Font(bold=True, size=11))
            row += 1

            for finding in self.report_data.management_summary.key_findings:
                self._add_wrapped_cell(ws, row, 1, f"• {finding}")
                row += 1

            row += 1

            # Alerts
            if self.report_data.management_summary.alerts:
                self._add_cell(ws, row, 1, "Alerts", Font(bold=True, size=11, color="FF0000"))
                row += 1

                for alert in self.report_data.management_summary.alerts:
                    self._add_wrapped_cell(ws, row, 1, f"⚠ {alert}")
                    row += 1

            row += 1

            # Opportunities
            if self.report_data.management_summary.opportunities:
                self._add_cell(ws, row, 1, "Opportunities", Font(bold=True, size=11, color="00B050"))
                row += 1

                for opp in self.report_data.management_summary.opportunities:
                    self._add_wrapped_cell(ws, row, 1, f"✓ {opp}")
                    row += 1

    def _create_platform_sheets(self):
        """Create one sheet per platform with performance breakdown."""
        for idx, platform in enumerate(self.report_data.platforms):
            sheet_name = platform.platform_name[:30]  # Excel sheet name limit
            ws = self.workbook.create_sheet(sheet_name)

            row = 1

            # Platform header
            header_fill = PatternFill(start_color="2c5aa0", end_color="2c5aa0", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)

            self._add_cell(ws, row, 1, f"{platform.platform_name} - Performance Metrics", header_font, header_fill)
            row += 2

            # KPI Summary
            kpi_data = [
                ("Gross Revenue", platform.gross_revenue),
                ("Returns & Refunds", platform.returns_refunds),
                ("Returns %", platform.returns_percentage),
                ("Net Realized Revenue", platform.net_revenue),
                ("Fulfillment OTIF %", platform.fulfillment_otif),
                ("Ad Spend", platform.ad_spend),
                ("Net Ad Cost", platform.net_ad_cost),
                ("TACoS %", platform.tacos_percentage),
                ("Net Profit", platform.net_profit),
                ("Margin %", platform.margin_percentage),
            ]

            for label, value in kpi_data:
                self._add_cell(ws, row, 1, label, Font(bold=True))
                self._add_currency_cell(ws, row, 2, value)
                row += 1

            row += 1

            # Product breakdown
            self._add_cell(ws, row, 1, "Product-Level Breakdown", Font(bold=True, size=11))
            row += 1

            # Headers
            headers = [
                "SKU",
                "Product Name",
                "Units Sold",
                "Gross Revenue",
                "Returns (Count)",
                "Returns %",
                "Organic Units",
                "Paid Units",
                "Ad Spend",
                "Net Ad Cost",
                "TACoS %",
                "Net Profit",
                "Margin %",
            ]

            header_fill = PatternFill(start_color="4a7ab7", end_color="4a7ab7", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=9)

            for col_idx, header in enumerate(headers, 1):
                self._add_cell(ws, row, col_idx, header, header_font, header_fill)

            row += 1

            # Product data
            for product in platform.products:
                col = 1
                self._add_cell(ws, row, col, product.sku)
                col += 1
                self._add_cell(ws, row, col, product.product_name)
                col += 1
                self._add_cell(ws, row, col, product.units_sold)
                col += 1
                self._add_currency_cell(ws, row, col, product.gross_revenue)
                col += 1
                self._add_cell(ws, row, col, product.returns_count)
                col += 1
                self._add_percentage_cell(ws, row, col, product.returns_percentage)
                col += 1
                self._add_cell(ws, row, col, product.organic_units)
                col += 1
                self._add_cell(ws, row, col, product.paid_units)
                col += 1
                self._add_currency_cell(ws, row, col, product.ad_spend)
                col += 1
                self._add_currency_cell(ws, row, col, product.net_ad_cost)
                col += 1
                self._add_percentage_cell(ws, row, col, product.tacos_percentage)
                col += 1
                self._add_currency_cell(ws, row, col, product.net_profit)
                col += 1
                self._add_percentage_cell(ws, row, col, product.margin_percentage)

                row += 1

            # Auto-adjust columns
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_consolidated_products_sheet(self):
        """Create consolidated product analysis sheet."""
        ws = self.workbook.create_sheet("Consolidated Products", len(self.workbook.sheetnames))
        ws.column_dimensions["A"].width = 50

        row = 1

        # Headers
        headers = [
            "SKU",
            "Product Name",
            "All Units",
            "Total Gross",
            "All Returns",
            "Returns %",
            "Org/Paid Split",
            "Total Ad Cost",
            "Net Ad Cost",
            "TACoS %",
            "Net Profit",
            "Margin %",
            "Stock DOS",
        ]

        header_fill = PatternFill(start_color="2c5aa0", end_color="2c5aa0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)

        for col_idx, header in enumerate(headers, 1):
            self._add_cell(ws, row, col_idx, header, header_font, header_fill)

        row += 1

        # Product data
        for product in self.report_data.consolidated_products:
            col = 1
            self._add_cell(ws, row, col, product.sku)
            col += 1
            self._add_cell(ws, row, col, product.product_name)
            col += 1
            self._add_cell(ws, row, col, product.all_units)
            col += 1
            self._add_currency_cell(ws, row, col, product.total_gross)
            col += 1
            self._add_cell(ws, row, col, product.all_returns)
            col += 1
            self._add_percentage_cell(ws, row, col, product.returns_percentage)
            col += 1
            self._add_cell(ws, row, col, product.organic_paid_split)
            col += 1
            self._add_currency_cell(ws, row, col, product.total_ad_cost)
            col += 1
            self._add_currency_cell(ws, row, col, product.net_ad_cost)
            col += 1
            self._add_percentage_cell(ws, row, col, product.tacos_percentage)
            col += 1
            self._add_currency_cell(ws, row, col, product.net_profit)
            col += 1
            self._add_percentage_cell(ws, row, col, product.margin_percentage)
            col += 1
            self._add_cell(ws, row, col, f"{product.stock_dos:.1f}")

            row += 1

    def _create_pnl_sheet(self):
        """Create P&L statement sheet."""
        ws = self.workbook.create_sheet("P&L Statement")
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15

        row = 1

        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2c5aa0", end_color="2c5aa0", fill_type="solid")

        self._add_cell(ws, row, 1, "P&L Statement", header_font, header_fill)
        self._add_cell(ws, row, 2, "Amount (INR)", header_font, header_fill)
        self._add_cell(ws, row, 3, "% of Revenue", header_font, header_fill)

        row += 1

        if self.report_data.pnl:
            pnl_items = [
                ("Total Gross Sales Turnover (GMV)", self.report_data.pnl.total_gross_gmv, Decimal("100")),
                ("Less: Returns & Customer Refunds", -self.report_data.pnl.less_returns_refunds, -self.report_data.pnl.less_returns_percentage),
                ("Net Realized Sales Turnover", self.report_data.pnl.net_revenue, Decimal("100")),
                ("Less: Cost of Goods Sold (COGS)", -self.report_data.pnl.less_cogs, -self.report_data.pnl.less_cogs_percentage),
                ("Less: Total Marketing & Ad Spend", -self.report_data.pnl.less_ad_spend, -self.report_data.pnl.less_ad_spend_percentage),
                ("Less: Marketplace Commission & Logistics", -self.report_data.pnl.less_commission_logistics, -self.report_data.pnl.less_commission_logistics_percentage),
                ("GRAND NET OPERATING PROFIT (EBITDA)", self.report_data.pnl.grand_net_operating_profit, self.report_data.pnl.margin_percentage),
            ]

            for label, amount, pct in pnl_items:
                bold = "GRAND" in label or "Net Realized" in label
                font = Font(bold=bold)
                self._add_cell(ws, row, 1, label, font)
                self._add_currency_cell(ws, row, 2, amount)
                self._add_percentage_cell(ws, row, 3, pct)
                row += 1

    def _create_channel_efficiency_sheet(self):
        """Create channel efficiency ranking sheet."""
        ws = self.workbook.create_sheet("Channel Efficiency")

        row = 1

        # Headers
        headers = [
            "Rank",
            "Platform",
            "Orders",
            "Units",
            "Gross Sales",
            "Sales Share %",
            "Ad Cost",
            "TACoS %",
            "Net Profit",
            "Margin %",
            "OTIF %",
        ]

        header_fill = PatternFill(start_color="2c5aa0", end_color="2c5aa0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)

        for col_idx, header in enumerate(headers, 1):
            self._add_cell(ws, row, col_idx, header, header_font, header_fill)

        row += 1

        # Channel data
        for channel in self.report_data.channel_efficiency:
            col = 1
            self._add_cell(ws, row, col, channel.rank)
            col += 1
            self._add_cell(ws, row, col, channel.platform_name)
            col += 1
            self._add_cell(ws, row, col, channel.orders)
            col += 1
            self._add_cell(ws, row, col, channel.units)
            col += 1
            self._add_currency_cell(ws, row, col, channel.gross_sales)
            col += 1
            self._add_percentage_cell(ws, row, col, channel.sales_share_percentage)
            col += 1
            self._add_currency_cell(ws, row, col, channel.ad_cost)
            col += 1
            self._add_percentage_cell(ws, row, col, channel.tacos_percentage)
            col += 1
            self._add_currency_cell(ws, row, col, channel.net_profit)
            col += 1
            self._add_percentage_cell(ws, row, col, channel.margin_percentage)
            col += 1
            self._add_percentage_cell(ws, row, col, channel.otif_percentage)

            row += 1

    def _create_recommendations_sheet(self):
        """Create recommendations and alerts sheet."""
        ws = self.workbook.create_sheet("Recommendations")
        ws.column_dimensions["A"].width = 100

        row = 1

        if self.report_data.management_summary:
            # Recommendations
            self._add_cell(ws, row, 1, "Recommendations", Font(bold=True, size=12))
            row += 1

            for idx, rec in enumerate(self.report_data.management_summary.recommendations, 1):
                self._add_wrapped_cell(ws, row, 1, f"{idx}. {rec}")
                row += 1

            row += 2

            # Alerts
            if self.report_data.management_summary.alerts:
                alert_font = Font(bold=True, size=11, color="FF0000")
                self._add_cell(ws, row, 1, "Alerts", alert_font)
                row += 1

                for alert in self.report_data.management_summary.alerts:
                    self._add_wrapped_cell(ws, row, 1, f"⚠ {alert}")
                    row += 1

            row += 2

            # Opportunities
            if self.report_data.management_summary.opportunities:
                opp_font = Font(bold=True, size=11, color="00B050")
                self._add_cell(ws, row, 1, "Opportunities", opp_font)
                row += 1

                for opp in self.report_data.management_summary.opportunities:
                    self._add_wrapped_cell(ws, row, 1, f"✓ {opp}")
                    row += 1

    def _add_cell(self, ws, row: int, col: int, value, font=None, fill=None):
        """Add a cell with optional formatting."""
        cell = ws.cell(row=row, column=col, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(wrap_text=False, vertical="center")

    def _add_wrapped_cell(self, ws, row: int, col: int, value):
        """Add a cell with text wrapping."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(30, len(value) // 30 * 15)

    def _add_currency_cell(self, ws, row: int, col: int, value):
        """Add a currency-formatted cell."""
        cell = ws.cell(row=row, column=col, value=float(value))
        cell.number_format = '"₹"#,##0'
        cell.alignment = Alignment(horizontal="right")

    def _add_percentage_cell(self, ws, row: int, col: int, value):
        """Add a percentage-formatted cell."""
        cell = ws.cell(row=row, column=col, value=float(value))
        cell.number_format = '0.00"%"'
        cell.alignment = Alignment(horizontal="center")

    def _generate_fallback(self) -> bytes:
        """Generate simple CSV if openpyxl is not available."""
        lines = [
            "SLEEPSIA AUDIT REPORT",
            f"Audit Date: {self.report_data.metadata.audit_date.strftime('%d %b %Y')}",
            f"Organization: {self.report_data.metadata.organization}",
            "",
            "P&L SUMMARY",
        ]

        if self.report_data.pnl:
            lines.append(f"Gross Revenue,{self.report_data.pnl.total_gross_gmv}")
            lines.append(f"Net Revenue,{self.report_data.pnl.net_revenue}")
            lines.append(f"Net Profit,{self.report_data.pnl.grand_net_operating_profit}")

        return "\n".join(lines).encode("utf-8")
